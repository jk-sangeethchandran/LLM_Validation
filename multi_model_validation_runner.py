#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
multi_model_validation_runner.py

Unified FSP validation runner for multiple instructor and target models.

What it does:
1. Lets you choose one or more instructor models and one or more target models before any data is sent.
2. Uses the official IPIP-NEO-300 key embedded from your "Conversational And Non-verbal.py" script.
3. The instructor model converts each participant's full 300-item questionnaire profile into a detailed,
   facet-by-facet personality control prompt.
4. The target model receives:
      - the instructor-generated personality prompt
      - calibration statistics from the selected training split
      - labelled training-item examples
   Then it answers the selected held-out validation items.
   Prompt Type 2 uses the FPS prompt path and now allows either questionnaire split: 120-180 or 180-120.
5. Runs all selected instructor × target combinations.
6. Caches instructor prompts, system prompts, target batches, and OpenAI prompt-cache hints where supported.
7. Compares human vs LLM answers after reverse coding both sides for trait-direction metrics.
8. Writes partial_results.jsonl after every successful case so interrupted runs can be resumed and merged.
9. Exports alignment-focused and mean-deviation Excel/CSV/JSONL results plus publication-friendly plots.
10. Creates Alignment Scores, Mean Dev, Combo Compare, and Plots sheets.

Install in the same venv:
    python -m pip install -U pandas numpy scipy matplotlib openpyxl python-dotenv groq openai "openai[realtime]" google-genai

Expected files beside this script, unless overridden:
    .env
    IPIP_NEO_300.csv   OR   prosocial_antisocial_ipip300_answers.csv

Relevant .env variables:
    OPENAI_API_KEY=...
    XAI_API_KEY=...                  # xAI/Grok API key, only needed for Grok models
    GROQ_API_KEY=...                 # Groq key 1, only needed if using llama-3.1-8b-instant
    GROQ_API_KEY_2=...               # optional Groq fallback key
    GROQ_API_KEY_3=...               # optional Groq fallback key
    GROQ_API_KEY_TUD=...             # optional Groq fallback key
    GEMINI_API_KEY=...               # Google AI Studio / Gemini API key, only needed for Gemma
    XAI_API_KEY=...                  # official xAI env var for Grok models
    GOOGLE_AI_STUDIO_API_KEY=...     # optional alternative env name
    GOOGLE_API_KEY=...               # optional alternative env name

Optional Groq throttling controls:
    GROQ_TPM_LIMIT=6000
    GROQ_TPM_SAFETY=0.95
    GROQ_MIN_INTERVAL_SECONDS=1
    GROQ_TOKEN_BUFFER=150
    GROQ_THROTTLE_ENABLED=1
    GROQ_KEY_ROTATE_AFTER_SECONDS=120   # rotate Groq key if retry hint is >= this long
    GROQ_SKIP_OVERSIZE_SLEEP=1           # do not sleep 60s for single requests above current TPM limit
    GROQ_TYPE2_INSTRUCTOR_MAX_OUTPUT_TOKENS=400  # exact FPS script used max_tokens=400 for Llama narrative

Gemma models are called through Google AI Studio / Gemini API using the google-genai SDK.
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import io
import json
import logging
import math
import os
import random
import re
import sys
import time
import threading
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from dotenv import load_dotenv

import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from scipy.stats import pearsonr

try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
except Exception:
    load_workbook = None
    XLImage = None


# =============================================================================
# Paths and runtime constants
# =============================================================================

SCRIPT_DIR = Path(__file__).resolve().parent
ENV_FILE = SCRIPT_DIR / ".env"
load_dotenv(dotenv_path=ENV_FILE, override=True)

PROMPT_VERSION = "multi_model_validation_v25_llama_target_parser_batchfix"
SELECTION_FILE = SCRIPT_DIR / ".validation_model_selection.json"
CACHE_DIR = SCRIPT_DIR / ".validation_cache"
OUTPUT_ROOT = SCRIPT_DIR / "validation_outputs"
RESUME_STATE_FILE = OUTPUT_ROOT / "_latest_resume_state.json"
PARTIAL_RESULTS_FILENAME = "partial_results.jsonl"
RUN_MANIFEST_FILENAME = "run_manifest.json"
CUMULATIVE_COMPARISON_CSV = OUTPUT_ROOT / "combo_summary_cumulative.csv"
CACHE_DIR.mkdir(exist_ok=True)
OUTPUT_ROOT.mkdir(exist_ok=True)

DEFAULT_MAX_PARTICIPANTS = 10
DEFAULT_BATCH_SIZE = 30
# Llama/Groq as TARGET can exceed the 6000 TPM tier when the persona prompt is long.
# Smaller target batches keep each single Groq request below the on-demand TPM limit.
LLAMA_TARGET_BATCH_SIZE = int(os.getenv("LLAMA_TARGET_BATCH_SIZE", "20"))
DEFAULT_REASONING_EFFORT = os.getenv("OPENAI_REASONING_EFFORT", "none").strip() or "none"
DEFAULT_OPENAI_MAX_OUTPUT_TOKENS = int(os.getenv("OPENAI_MAX_OUTPUT_TOKENS", "1800"))
DEFAULT_INSTRUCTOR_MAX_TOKENS = int(os.getenv("INSTRUCTOR_MAX_OUTPUT_TOKENS", "1800"))
DEFAULT_GEMINI_MAX_OUTPUT_TOKENS = int(os.getenv("GEMINI_MAX_OUTPUT_TOKENS", "1800"))
DEFAULT_GROQ_MAX_TOKENS = int(os.getenv("GROQ_MAX_OUTPUT_TOKENS", "1800"))

# Groq/Llama on-demand tiers often fail because of TPM rather than context length.
# The defaults below are deliberately conservative for the common 6000 TPM on-demand limit.
# You can tune them in .env if your Groq account has a higher limit.
GROQ_THROTTLE_ENABLED = (os.getenv("GROQ_THROTTLE_ENABLED", "1").strip() != "0")
GROQ_TPM_LIMIT = int(os.getenv("GROQ_TPM_LIMIT", "6000"))
GROQ_TPM_SAFETY = float(os.getenv("GROQ_TPM_SAFETY", "0.95"))
GROQ_MIN_INTERVAL_SECONDS = float(os.getenv("GROQ_MIN_INTERVAL_SECONDS", "1.0"))
GROQ_EST_CHARS_PER_TOKEN = float(os.getenv("GROQ_EST_CHARS_PER_TOKEN", "4.0"))
GROQ_TOKEN_BUFFER = int(os.getenv("GROQ_TOKEN_BUFFER", "150"))
GROQ_KEY_ROTATE_AFTER_SECONDS = float(os.getenv("GROQ_KEY_ROTATE_AFTER_SECONDS", "120"))
GROQ_SKIP_OVERSIZE_SLEEP = (os.getenv("GROQ_SKIP_OVERSIZE_SLEEP", "1").strip() != "0")
GROQ_TYPE2_INSTRUCTOR_MAX_OUTPUT_TOKENS = int(os.getenv("GROQ_TYPE2_INSTRUCTOR_MAX_OUTPUT_TOKENS", "400"))

OPENAI_API_KEY = (os.getenv("OPENAI_API_KEY") or "").strip()
XAI_API_KEY = (os.getenv("XAI_API_KEY") or os.getenv("GROK_API_KEY") or "").strip()
XAI_BASE_URL = (os.getenv("XAI_BASE_URL") or "https://api.x.ai/v1").strip()

# Groq key rotation order requested for long daily/quota waits.
# The runner starts with the first available key and switches at runtime when Groq
# returns a long retry window, e.g., around 300s / 5 minutes.
GROQ_API_KEY_ENV_ORDER = ["GROQ_API_KEY", "GROQ_API_KEY_2", "GROQ_API_KEY_3", "GROQ_API_KEY_TUD"]
GROQ_API_KEY_ENTRIES = [
    (name, (os.getenv(name) or "").strip())
    for name in GROQ_API_KEY_ENV_ORDER
    if (os.getenv(name) or "").strip()
]
GROQ_API_KEY = GROQ_API_KEY_ENTRIES[0][1] if GROQ_API_KEY_ENTRIES else ""
GEMINI_API_KEY = (os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_AI_STUDIO_API_KEY") or os.getenv("GOOGLE_API_KEY") or "").strip()

TRAIN_ITEMS = list(range(1, 181))
TEST_ITEMS = list(range(181, 301))
ALL_ITEMS = list(range(1, 301))
TRAIN_COLS = [f"i{n}" for n in TRAIN_ITEMS]
TEST_COLS = [f"i{n}" for n in TEST_ITEMS]
ALL_COLS = [f"i{n}" for n in ALL_ITEMS]

# Questionnaire split options:
# - 120-180: target sees labelled items 1-120 and answers items 121-300
# - 180-120: target sees labelled items 1-180 and answers items 181-300
QUESTIONNAIRE_MODE = "180-120"
TRAIN_RANGE_LABEL = "items 1-180"
TEST_RANGE_LABEL = "items 181-300"
TEST_SCORE_SUFFIX = "120"

def split_config(mode: str) -> tuple[list[int], list[int]]:
    mode = str(mode or "120-180").strip()
    if mode == "180-120":
        return list(range(1, 181)), list(range(181, 301))
    return list(range(1, 121)), list(range(121, 301))

def questionnaire_mode_description(mode: str) -> str:
    if str(mode) == "180-120":
        return "180-120 — target sees labelled items 1-180 and answers items 181-300"
    return "120-180 — target sees labelled items 1-120 and answers items 121-300"

def configure_questionnaire_split(mode: str) -> str:
    """Update global train/test split variables before any participant is processed."""
    global QUESTIONNAIRE_MODE, TRAIN_ITEMS, TEST_ITEMS, TRAIN_COLS, TEST_COLS
    global TRAIN_KEY, TEST_KEY, TRAIN_DOMAIN_COUNTS, TEST_DOMAIN_COUNTS
    global TRAIN_RANGE_LABEL, TEST_RANGE_LABEL, TEST_SCORE_SUFFIX

    mode = "180-120" if str(mode).strip() == "180-120" else "120-180"
    train_items, test_items = split_config(mode)

    QUESTIONNAIRE_MODE = mode
    TRAIN_ITEMS = train_items
    TEST_ITEMS = test_items
    TRAIN_COLS = [f"i{n}" for n in TRAIN_ITEMS]
    TEST_COLS = [f"i{n}" for n in TEST_ITEMS]
    TRAIN_RANGE_LABEL = f"items {TRAIN_ITEMS[0]}-{TRAIN_ITEMS[-1]}"
    TEST_RANGE_LABEL = f"items {TEST_ITEMS[0]}-{TEST_ITEMS[-1]}"
    TEST_SCORE_SUFFIX = str(len(TEST_ITEMS))

    if "ITEM_KEY_300" in globals():
        train_set = set(TRAIN_ITEMS)
        test_set = set(TEST_ITEMS)
        TRAIN_KEY = [entry for entry in ITEM_KEY_300 if entry[0] in train_set]
        TEST_KEY = [entry for entry in ITEM_KEY_300 if entry[0] in test_set]
        TRAIN_DOMAIN_COUNTS = {
            d: sum(1 for _, domain, _, _, _ in TRAIN_KEY if domain == d)
            for d in DOMAIN_ORDER
        }
        TEST_DOMAIN_COUNTS = {
            d: sum(1 for _, domain, _, _, _ in TEST_KEY if domain == d)
            for d in DOMAIN_ORDER
        }
    return mode

DOMAIN_ORDER = ["O", "C", "E", "A", "N"]
DOMAIN_FULL_NAMES = {
    "O": "Openness",
    "C": "Conscientiousness",
    "E": "Extraversion",
    "A": "Agreeableness",
    "N": "Neuroticism",
}

LABEL_MAP_FORMAL = {
    1: "Very Inaccurate",
    2: "Moderately Inaccurate",
    3: "Neither Accurate Nor Inaccurate",
    4: "Moderately Accurate",
    5: "Very Accurate",
    0: "Neither Accurate Nor Inaccurate",
}

SCORE_LABEL_FOR_INSTRUCTOR = {
    1: "almost never / strongly inaccurate",
    2: "rarely / moderately inaccurate",
    3: "sometimes / neither accurate nor inaccurate",
    4: "often / moderately accurate",
    5: "almost always / very accurate",
}

# Embedded from your Conversational And Non-verbal.py official-key script.
# Tuple format: (item_number, domain, facet, reverse_scored, item_text)
ITEM_KEY_300 = [(1, 'N', 'Anxiety', False, 'Worry about things.'),
 (2, 'E', 'Friendliness', False, 'Make friends easily.'),
 (3, 'O', 'Imagination', False, 'Have a vivid imagination.'),
 (4, 'A', 'Trust', False, 'Trust others.'),
 (5, 'C', 'Self-Efficacy', False, 'Complete tasks successfully.'),
 (6, 'N', 'Anger', False, 'Get angry easily.'),
 (7, 'E', 'Gregariousness', False, 'Love large parties.'),
 (8, 'O', 'Artistic Interests', False, 'Believe in the importance of art.'),
 (9, 'A', 'Morality', False, 'Would never cheat on my taxes.'),
 (10, 'C', 'Orderliness', False, 'Like order.'),
 (11, 'N', 'Depression', False, 'Often feel blue.'),
 (12, 'E', 'Assertiveness', False, 'Take charge.'),
 (13, 'O', 'Emotionality', False, 'Experience my emotions intensely.'),
 (14, 'A', 'Altruism', False, 'Make people feel welcome.'),
 (15, 'C', 'Dutifulness', False, 'Try to follow the rules.'),
 (16, 'N', 'Self-Consciousness', False, 'Am easily intimidated.'),
 (17, 'E', 'Activity Level', False, 'Am always busy.'),
 (18, 'O', 'Adventurousness', False, 'Prefer variety to routine.'),
 (19, 'A', 'Cooperation', False, 'Am easy to satisfy.'),
 (20, 'C', 'Achievement-Striving', False, 'Go straight for the goal.'),
 (21, 'N', 'Immoderation', False, 'Often eat too much.'),
 (22, 'E', 'Excitement-Seeking', False, 'Love excitement.'),
 (23, 'O', 'Intellect', False, 'Like to solve complex problems.'),
 (24, 'A', 'Modesty', False, 'Dislike being the center of attention.'),
 (25, 'C', 'Self-Discipline', False, 'Get chores done right away.'),
 (26, 'N', 'Vulnerability', False, 'Panic easily.'),
 (27, 'E', 'Cheerfulness', False, 'Radiate joy.'),
 (28, 'O', 'Liberalism', False, 'Tend to vote for liberal political candidates.'),
 (29, 'A', 'Sympathy', False, 'Sympathize with the homeless.'),
 (30, 'C', 'Cautiousness', False, 'Avoid mistakes.'),
 (31, 'N', 'Anxiety', False, 'Fear for the worst.'),
 (32, 'E', 'Friendliness', False, 'Warm up quickly to others.'),
 (33, 'O', 'Imagination', False, 'Enjoy wild flights of fantasy.'),
 (34, 'A', 'Trust', False, 'Believe that others have good intentions.'),
 (35, 'C', 'Self-Efficacy', False, 'Excel in what I do.'),
 (36, 'N', 'Anger', False, 'Get irritated easily.'),
 (37, 'E', 'Gregariousness', False, 'Talk to a lot of different people at parties.'),
 (38, 'O', 'Artistic Interests', False, 'Like music.'),
 (39, 'A', 'Morality', False, 'Stick to the rules.'),
 (40, 'C', 'Orderliness', False, 'Like to tidy up.'),
 (41, 'N', 'Depression', False, 'Dislike myself.'),
 (42, 'E', 'Assertiveness', False, 'Try to lead others.'),
 (43, 'O', 'Emotionality', False, "Feel others' emotions."),
 (44, 'A', 'Altruism', False, 'Anticipate the needs of others.'),
 (45, 'C', 'Dutifulness', False, 'Keep my promises.'),
 (46, 'N', 'Self-Consciousness', False, 'Am afraid that I will do the wrong thing.'),
 (47, 'E', 'Activity Level', False, 'Am always on the go.'),
 (48, 'O', 'Adventurousness', False, 'Like to visit new places.'),
 (49, 'A', 'Cooperation', False, "Can't stand confrontations."),
 (50, 'C', 'Achievement-Striving', False, 'Work hard.'),
 (51, 'N', 'Immoderation', False, "Don't know why I do some of the things I do."),
 (52, 'E', 'Excitement-Seeking', False, 'Seek adventure.'),
 (53, 'O', 'Intellect', False, 'Love to read challenging material.'),
 (54, 'A', 'Modesty', False, 'Dislike talking about myself.'),
 (55, 'C', 'Self-Discipline', False, 'Am always prepared.'),
 (56, 'N', 'Vulnerability', False, 'Become overwhelmed by events.'),
 (57, 'E', 'Cheerfulness', False, 'Have a lot of fun.'),
 (58, 'O', 'Liberalism', False, 'Believe that there is no absolute right or wrong.'),
 (59, 'A', 'Sympathy', False, 'Feel sympathy for those who are worse off than myself.'),
 (60, 'C', 'Cautiousness', False, 'Choose my words with care.'),
 (61, 'N', 'Anxiety', False, 'Am afraid of many things.'),
 (62, 'E', 'Friendliness', False, 'Feel comfortable around people.'),
 (63, 'O', 'Imagination', False, 'Love to daydream.'),
 (64, 'A', 'Trust', False, 'Trust what people say.'),
 (65, 'C', 'Self-Efficacy', False, 'Handle tasks smoothly.'),
 (66, 'N', 'Anger', False, 'Get upset easily.'),
 (67, 'E', 'Gregariousness', False, 'Enjoy being part of a group.'),
 (68, 'O', 'Artistic Interests', False, 'See beauty in things that others might not notice.'),
 (69, 'A', 'Morality', True, 'Use flattery to get ahead.'),
 (70, 'C', 'Orderliness', False, 'Want everything to be "just right."'),
 (71, 'N', 'Depression', False, 'Am often down in the dumps.'),
 (72, 'E', 'Assertiveness', False, 'Can talk others into doing things.'),
 (73, 'O', 'Emotionality', False, 'Am passionate about causes.'),
 (74, 'A', 'Altruism', False, 'Love to help others.'),
 (75, 'C', 'Dutifulness', False, 'Pay my bills on time.'),
 (76, 'N', 'Self-Consciousness', False, 'Find it difficult to approach others.'),
 (77, 'E', 'Activity Level', False, 'Do a lot in my spare time.'),
 (78, 'O', 'Adventurousness', False, 'Interested in many things.'),
 (79, 'A', 'Cooperation', False, 'Hate to seem pushy.'),
 (80, 'C', 'Achievement-Striving', False, 'Turn plans into actions.'),
 (81, 'N', 'Immoderation', False, 'Do things I later regret.'),
 (82, 'E', 'Excitement-Seeking', False, 'Love action.'),
 (83, 'O', 'Intellect', False, 'Have a rich vocabulary.'),
 (84, 'A', 'Modesty', False, 'Consider myself an average person.'),
 (85, 'C', 'Self-Discipline', False, 'Start tasks right away.'),
 (86, 'N', 'Vulnerability', False, "Feel that I'm unable to deal with things."),
 (87, 'E', 'Cheerfulness', False, 'Express childlike joy.'),
 (88, 'O', 'Liberalism', False, 'Believe that criminals should receive help rather than punishment.'),
 (89, 'A', 'Sympathy', False, 'Value cooperation over competition.'),
 (90, 'C', 'Cautiousness', False, 'Stick to my chosen path.'),
 (91, 'N', 'Anxiety', False, 'Get stressed out easily.'),
 (92, 'E', 'Friendliness', False, 'Act comfortably with others.'),
 (93, 'O', 'Imagination', False, 'Like to get lost in thought.'),
 (94, 'A', 'Trust', False, 'Believe that people are basically moral.'),
 (95, 'C', 'Self-Efficacy', False, 'Am sure of my ground.'),
 (96, 'N', 'Anger', False, 'Am often in a bad mood.'),
 (97, 'E', 'Gregariousness', False, 'Involve others in what I am doing.'),
 (98, 'O', 'Artistic Interests', False, 'Love flowers.'),
 (99, 'A', 'Morality', True, 'Use others for my own ends.'),
 (100, 'C', 'Orderliness', False, 'Love order and regularity.'),
 (101, 'N', 'Depression', False, 'Have a low opinion of myself.'),
 (102, 'E', 'Assertiveness', False, 'Seek to influence others.'),
 (103, 'O', 'Emotionality', False, 'Enjoy examining myself and my life.'),
 (104, 'A', 'Altruism', False, 'Am concerned about others.'),
 (105, 'C', 'Dutifulness', False, 'Tell the truth.'),
 (106, 'N', 'Self-Consciousness', False, 'Am afraid to draw attention to myself.'),
 (107, 'E', 'Activity Level', False, 'Can manage many things at the same time.'),
 (108, 'O', 'Adventurousness', False, 'Like to begin new things.'),
 (109, 'A', 'Cooperation', True, 'Have a sharp tongue.'),
 (110, 'C', 'Achievement-Striving', False, 'Plunge into tasks with all my heart.'),
 (111, 'N', 'Immoderation', False, 'Go on binges.'),
 (112, 'E', 'Excitement-Seeking', False, 'Enjoy being part of a loud crowd.'),
 (113, 'O', 'Intellect', False, 'Can handle a lot of information.'),
 (114, 'A', 'Modesty', False, 'Seldom toot my own horn.'),
 (115, 'C', 'Self-Discipline', False, 'Get to work at once.'),
 (116, 'N', 'Vulnerability', False, "Can't make up my mind."),
 (117, 'E', 'Cheerfulness', False, 'Laugh my way through life.'),
 (118, 'O', 'Liberalism', True, 'Believe in one true religion.'),
 (119, 'A', 'Sympathy', False, "Suffer from others' sorrows."),
 (120, 'C', 'Cautiousness', True, 'Jump into things without thinking.'),
 (121, 'N', 'Anxiety', False, 'Get caught up in my problems.'),
 (122, 'E', 'Friendliness', False, 'Cheer people up.'),
 (123, 'O', 'Imagination', False, 'Indulge in my fantasies.'),
 (124, 'A', 'Trust', False, 'Believe in human goodness.'),
 (125, 'C', 'Self-Efficacy', False, 'Come up with good solutions.'),
 (126, 'N', 'Anger', False, 'Lose my temper.'),
 (127, 'E', 'Gregariousness', False, 'Love surprise parties.'),
 (128, 'O', 'Artistic Interests', False, 'Enjoy the beauty of nature.'),
 (129, 'A', 'Morality', True, 'Know how to get around the rules.'),
 (130, 'C', 'Orderliness', False, 'Do things according to a plan.'),
 (131, 'N', 'Depression', False, 'Have frequent mood swings.'),
 (132, 'E', 'Assertiveness', False, 'Take control of things.'),
 (133, 'O', 'Emotionality', False, 'Try to understand myself.'),
 (134, 'A', 'Altruism', False, 'Have a good word for everyone.'),
 (135, 'C', 'Dutifulness', False, 'Listen to my conscience.'),
 (136, 'N', 'Self-Consciousness', False, 'Only feel comfortable with friends.'),
 (137, 'E', 'Activity Level', False, 'React quickly.'),
 (138, 'O', 'Adventurousness', True, 'Prefer to stick with things that I know.'),
 (139, 'A', 'Cooperation', True, 'Contradict others.'),
 (140, 'C', 'Achievement-Striving', False, "Do more than what's expected of me."),
 (141, 'N', 'Immoderation', False, 'Love to eat.'),
 (142, 'E', 'Excitement-Seeking', False, 'Enjoy being reckless.'),
 (143, 'O', 'Intellect', False, 'Enjoy thinking about things.'),
 (144, 'A', 'Modesty', True, 'Believe that I am better than others.'),
 (145, 'C', 'Self-Discipline', False, 'Carry out my plans.'),
 (146, 'N', 'Vulnerability', False, 'Get overwhelmed by emotions.'),
 (147, 'E', 'Cheerfulness', False, 'Love life.'),
 (148, 'O', 'Liberalism', True, 'Tend to vote for conservative political candidates.'),
 (149, 'A', 'Sympathy', True, "Am not interested in other people's problems."),
 (150, 'C', 'Cautiousness', True, 'Make rash decisions.'),
 (151, 'N', 'Anxiety', True, 'Am not easily bothered by things.'),
 (152, 'E', 'Friendliness', True, 'Am hard to get to know.'),
 (153, 'O', 'Imagination', False, 'Spend time reflecting on things.'),
 (154, 'A', 'Trust', False, 'Think that all will be well.'),
 (155, 'C', 'Self-Efficacy', False, 'Know how to get things done.'),
 (156, 'N', 'Anger', True, 'Rarely get irritated.'),
 (157, 'E', 'Gregariousness', True, 'Prefer to be alone.'),
 (158, 'O', 'Artistic Interests', True, 'Do not like art.'),
 (159, 'A', 'Morality', True, 'Cheat to get ahead.'),
 (160, 'C', 'Orderliness', True, 'Often forget to put things back in their proper place.'),
 (161, 'N', 'Depression', False, 'Feel desperate.'),
 (162, 'E', 'Assertiveness', True, 'Wait for others to lead the way.'),
 (163, 'O', 'Emotionality', True, 'Seldom get emotional.'),
 (164, 'A', 'Altruism', True, 'Look down on others.'),
 (165, 'C', 'Dutifulness', True, 'Break rules.'),
 (166, 'N', 'Self-Consciousness', False, 'Stumble over my words.'),
 (167, 'E', 'Activity Level', True, 'Like to take it easy.'),
 (168, 'O', 'Adventurousness', True, 'Dislike changes.'),
 (169, 'A', 'Cooperation', True, 'Love a good fight.'),
 (170, 'C', 'Achievement-Striving', False, 'Set high standards for myself and others.'),
 (171, 'N', 'Immoderation', True, 'Rarely overindulge.'),
 (172, 'E', 'Excitement-Seeking', False, 'Act wild and crazy.'),
 (173, 'O', 'Intellect', True, 'Am not interested in abstract ideas.'),
 (174, 'A', 'Modesty', True, 'Think highly of myself.'),
 (175, 'C', 'Self-Discipline', True, 'Find it difficult to get down to work.'),
 (176, 'N', 'Vulnerability', True, 'Remain calm under pressure.'),
 (177, 'E', 'Cheerfulness', False, 'Look at the bright side of life.'),
 (178, 'O', 'Liberalism', True, 'Believe that too much tax money goes to support artists.'),
 (179, 'A', 'Sympathy', True, 'Tend to dislike soft-hearted people.'),
 (180, 'C', 'Cautiousness', True, 'Like to act on a whim.'),
 (181, 'N', 'Anxiety', True, 'Am relaxed most of the time.'),
 (182, 'E', 'Friendliness', True, 'Often feel uncomfortable around others.'),
 (183, 'O', 'Imagination', True, 'Seldom daydream.'),
 (184, 'A', 'Trust', True, 'Distrust people.'),
 (185, 'C', 'Self-Efficacy', True, 'Misjudge situations.'),
 (186, 'N', 'Anger', True, 'Seldom get mad.'),
 (187, 'E', 'Gregariousness', True, 'Want to be left alone.'),
 (188, 'O', 'Artistic Interests', True, 'Do not like poetry.'),
 (189, 'A', 'Morality', True, 'Put people under pressure.'),
 (190, 'C', 'Orderliness', True, 'Leave a mess in my room.'),
 (191, 'N', 'Depression', False, 'Feel that my life lacks direction.'),
 (192, 'E', 'Assertiveness', True, 'Keep in the background.'),
 (193, 'O', 'Emotionality', True, 'Am not easily affected by my emotions.'),
 (194, 'A', 'Altruism', True, 'Am indifferent to the feelings of others.'),
 (195, 'C', 'Dutifulness', True, 'Break my promises.'),
 (196, 'N', 'Self-Consciousness', True, 'Am not embarrassed easily.'),
 (197, 'E', 'Activity Level', True, 'Like to take my time.'),
 (198, 'O', 'Adventurousness', True, "Don't like the idea of change."),
 (199, 'A', 'Cooperation', True, 'Yell at people.'),
 (200, 'C', 'Achievement-Striving', False, 'Demand quality.'),
 (201, 'N', 'Immoderation', True, 'Easily resist temptations.'),
 (202, 'E', 'Excitement-Seeking', False, 'Willing to try anything once.'),
 (203, 'O', 'Intellect', True, 'Avoid philosophical discussions.'),
 (204, 'A', 'Modesty', True, 'Have a high opinion of myself.'),
 (205, 'C', 'Self-Discipline', True, 'Waste my time.'),
 (206, 'N', 'Vulnerability', True, 'Can handle complex problems.'),
 (207, 'E', 'Cheerfulness', False, 'Laugh aloud.'),
 (208, 'O', 'Liberalism', True, 'Believe laws should be strictly enforced.'),
 (209, 'A', 'Sympathy', True, 'Believe in an eye for an eye.'),
 (210, 'C', 'Cautiousness', True, 'Rush into things.'),
 (211, 'N', 'Anxiety', True, 'Am not easily disturbed by events.'),
 (212, 'E', 'Friendliness', True, 'Avoid contacts with others.'),
 (213, 'O', 'Imagination', True, 'Do not have a good imagination.'),
 (214, 'A', 'Trust', True, 'Suspect hidden motives in others.'),
 (215, 'C', 'Self-Efficacy', True, "Don't understand things."),
 (216, 'N', 'Anger', True, 'Am not easily annoyed.'),
 (217, 'E', 'Gregariousness', True, "Don't like crowded events."),
 (218, 'O', 'Artistic Interests', True, 'Do not enjoy going to art museums.'),
 (219, 'A', 'Morality', True, 'Pretend to be concerned for others.'),
 (220, 'C', 'Orderliness', True, 'Leave my belongings around.'),
 (221, 'N', 'Depression', True, 'Seldom feel blue.'),
 (222, 'E', 'Assertiveness', True, 'Have little to say.'),
 (223, 'O', 'Emotionality', True, 'Rarely notice my emotional reactions.'),
 (224, 'A', 'Altruism', True, 'Make people feel uncomfortable.'),
 (225, 'C', 'Dutifulness', True, 'Get others to do my duties.'),
 (226, 'N', 'Self-Consciousness', True, 'Am comfortable in unfamiliar situations.'),
 (227, 'E', 'Activity Level', True, 'Like a leisurely lifestyle.'),
 (228, 'O', 'Adventurousness', True, 'Am a creature of habit.'),
 (229, 'A', 'Cooperation', True, 'Insult people.'),
 (230, 'C', 'Achievement-Striving', True, 'Am not highly motivated to succeed.'),
 (231, 'N', 'Immoderation', True, 'Am able to control my cravings.'),
 (232, 'E', 'Excitement-Seeking', False, 'Seek danger.'),
 (233, 'O', 'Intellect', True, 'Have difficulty understanding abstract ideas.'),
 (234, 'A', 'Modesty', True, 'Know the answers to many questions.'),
 (235, 'C', 'Self-Discipline', True, 'Need a push to get started.'),
 (236, 'N', 'Vulnerability', True, 'Know how to cope.'),
 (237, 'E', 'Cheerfulness', False, 'Amuse my friends.'),
 (238, 'O', 'Liberalism', True, 'Believe that we coddle criminals too much.'),
 (239, 'A', 'Sympathy', True, 'Try not to think about the needy.'),
 (240, 'C', 'Cautiousness', True, 'Do crazy things.'),
 (241, 'N', 'Anxiety', True, "Don't worry about things that have already happened."),
 (242, 'E', 'Friendliness', True, 'Am not really interested in others.'),
 (243, 'O', 'Imagination', True, 'Seldom get lost in thought.'),
 (244, 'A', 'Trust', True, 'Am wary of others.'),
 (245, 'C', 'Self-Efficacy', True, 'Have little to contribute.'),
 (246, 'N', 'Anger', True, 'Keep my cool.'),
 (247, 'E', 'Gregariousness', True, 'Avoid crowds.'),
 (248, 'O', 'Artistic Interests', True, 'Do not like concerts.'),
 (249, 'A', 'Morality', True, 'Take advantage of others.'),
 (250, 'C', 'Orderliness', True, 'Am not bothered by messy people.'),
 (251, 'N', 'Depression', True, 'Feel comfortable with myself.'),
 (252, 'E', 'Assertiveness', True, "Don't like to draw attention to myself."),
 (253, 'O', 'Emotionality', True, 'Experience very few emotional highs and lows.'),
 (254, 'A', 'Altruism', True, 'Turn my back on others.'),
 (255, 'C', 'Dutifulness', True, 'Do the opposite of what is asked.'),
 (256, 'N', 'Self-Consciousness', True, 'Am not bothered by difficult social situations.'),
 (257, 'E', 'Activity Level', True, 'Let things proceed at their own pace.'),
 (258, 'O', 'Adventurousness', True, 'Dislike new foods.'),
 (259, 'A', 'Cooperation', True, 'Get back at others.'),
 (260, 'C', 'Achievement-Striving', True, 'Do just enough work to get by.'),
 (261, 'N', 'Immoderation', True, 'Never spend more than I can afford.'),
 (262, 'E', 'Excitement-Seeking', True, 'Would never go hang gliding or bungee jumping.'),
 (263, 'O', 'Intellect', True, 'Am not interested in theoretical discussions.'),
 (264, 'A', 'Modesty', True, 'Boast about my virtues.'),
 (265, 'C', 'Self-Discipline', True, 'Have difficulty starting tasks.'),
 (266, 'N', 'Vulnerability', True, 'Readily overcome setbacks.'),
 (267, 'E', 'Cheerfulness', True, 'Am not easily amused.'),
 (268, 'O', 'Liberalism', True, 'Believe that we should be tough on crime.'),
 (269, 'A', 'Sympathy', True, 'Believe people should fend for themselves.'),
 (270, 'C', 'Cautiousness', True, 'Act without thinking.'),
 (271, 'N', 'Anxiety', True, 'Adapt easily to new situations.'),
 (272, 'E', 'Friendliness', True, 'Keep others at a distance.'),
 (273, 'O', 'Imagination', True, 'Have difficulty imagining things.'),
 (274, 'A', 'Trust', True, 'Believe that people are essentially evil.'),
 (275, 'C', 'Self-Efficacy', True, "Don't see the consequences of things."),
 (276, 'N', 'Anger', True, 'Rarely complain.'),
 (277, 'E', 'Gregariousness', True, 'Seek quiet.'),
 (278, 'O', 'Artistic Interests', True, 'Do not enjoy watching dance performances.'),
 (279, 'A', 'Morality', True, "Obstruct others' plans."),
 (280, 'C', 'Orderliness', True, 'Am not bothered by disorder.'),
 (281, 'N', 'Depression', True, 'Am very pleased with myself.'),
 (282, 'E', 'Assertiveness', True, 'Hold back my opinions.'),
 (283, 'O', 'Emotionality', True, "Don't understand people who get emotional."),
 (284, 'A', 'Altruism', True, 'Take no time for others.'),
 (285, 'C', 'Dutifulness', True, 'Misrepresent the facts.'),
 (286, 'N', 'Self-Consciousness', True, 'Am able to stand up for myself.'),
 (287, 'E', 'Activity Level', True, 'React slowly.'),
 (288, 'O', 'Adventurousness', True, 'Am attached to conventional ways.'),
 (289, 'A', 'Cooperation', True, 'Hold a grudge.'),
 (290, 'C', 'Achievement-Striving', True, 'Put little time and effort into my work.'),
 (291, 'N', 'Immoderation', True, 'Never splurge.'),
 (292, 'E', 'Excitement-Seeking', True, 'Dislike loud music.'),
 (293, 'O', 'Intellect', True, 'Avoid difficult reading material.'),
 (294, 'A', 'Modesty', True, 'Make myself the center of attention.'),
 (295, 'C', 'Self-Discipline', True, 'Postpone decisions.'),
 (296, 'N', 'Vulnerability', True, 'Am calm even in tense situations.'),
 (297, 'E', 'Cheerfulness', True, 'Seldom joke around.'),
 (298, 'O', 'Liberalism', True, 'Like to stand during the national anthem.'),
 (299, 'A', 'Sympathy', True, "Can't stand weak people."),
 (300, 'C', 'Cautiousness', True, 'Often make last-minute plans.')]

ITEM_TO_DOMAIN = {item_num: domain for item_num, domain, facet, reverse, item_text in ITEM_KEY_300}
ITEM_TO_FACET = {item_num: facet for item_num, domain, facet, reverse, item_text in ITEM_KEY_300}
ITEM_TO_REVERSE = {item_num: reverse for item_num, domain, facet, reverse, item_text in ITEM_KEY_300}
ITEM_TO_TEXT = {item_num: item_text for item_num, domain, facet, reverse, item_text in ITEM_KEY_300}
FACET_ORDER = []
FACET_TO_DOMAIN = {}
for item_num, domain, facet, reverse, item_text in ITEM_KEY_300:
    if facet not in FACET_ORDER:
        FACET_ORDER.append(facet)
    FACET_TO_DOMAIN[facet] = domain

# Default runtime split is overwritten after the interactive questionnaire-mode selection.
configure_questionnaire_split(QUESTIONNAIRE_MODE)
FULL_DOMAIN_COUNTS = {d: sum(1 for item_num, domain, facet, reverse, text in ITEM_KEY_300 if domain == d) for d in DOMAIN_ORDER}
FULL_FACET_COUNTS = {f: sum(1 for item_num, domain, facet, reverse, text in ITEM_KEY_300 if facet == f) for f in FACET_ORDER}

assert len(ITEM_KEY_300) == 300, "ITEM_KEY_300 must contain exactly 300 items."


# =============================================================================
# Model registry
# =============================================================================

@dataclass(frozen=True)
class ModelSpec:
    key: str
    display: str
    provider: str
    model_id: str
    notes: str


AVAILABLE_MODELS: Dict[str, ModelSpec] = {
    "gpt-5.4": ModelSpec(
        key="gpt-5.4",
        display="GPT-5.4",
        provider="openai_responses",
        model_id="gpt-5.4",
        notes="OpenAI Responses API",
    ),
    "gpt-5.4-mini": ModelSpec(
        key="gpt-5.4-mini",
        display="GPT-5.4 mini",
        provider="openai_responses",
        model_id="gpt-5.4-mini",
        notes="OpenAI Responses API",
    ),
    "gpt-5.4-nano": ModelSpec(
        key="gpt-5.4-nano",
        display="GPT-5.4 nano",
        provider="openai_responses",
        model_id="gpt-5.4-nano",
        notes="OpenAI Responses API",
    ),
    "gpt-realtime-1.5": ModelSpec(
        key="gpt-realtime-1.5",
        display="GPT-Realtime-1.5",
        provider="openai_realtime",
        model_id="gpt-realtime-1.5",
        notes="OpenAI Realtime API, text-only validation mode",
    ),
    "gemma-4-26b-a4b-it": ModelSpec(
        key="gemma-4-26b-a4b-it",
        display="Gemma 4 26B A4B IT",
        provider="google_genai",
        model_id="gemma-4-26b-a4b-it",
        notes="Google AI Studio / Gemini API via google-genai",
    ),
    "gemma-4-31b-it": ModelSpec(
        key="gemma-4-31b-it",
        display="Gemma 4 31B IT",
        provider="google_genai",
        model_id="gemma-4-31b-it",
        notes="Google AI Studio / Gemini API via google-genai",
    ),
    "gemini-2.5-flash": ModelSpec(
        key="gemini-2.5-flash",
        display="Gemini 2.5 Flash",
        provider="google_genai",
        model_id="gemini-2.5-flash",
        notes="Google AI Studio / Gemini API; faster target/instructor option",
    ),
    "gemini-2.5-flash-lite": ModelSpec(
        key="gemini-2.5-flash-lite",
        display="Gemini 2.5 Flash-Lite",
        provider="google_genai",
        model_id="gemini-2.5-flash-lite",
        notes="Google AI Studio / Gemini API; fastest/cheapest target option",
    ),
    "grok-4.20-reasoning": ModelSpec(
        key="grok-4.20-reasoning",
        display="Grok 4.20 Reasoning",
        provider="xai_responses",
        model_id="grok-4.20-reasoning",
        notes="xAI Responses API; official base URL https://api.x.ai/v1",
    ),
    "grok-4-1-fast-reasoning": ModelSpec(
        key="grok-4-1-fast-reasoning",
        display="Grok 4.1 Fast Reasoning",
        provider="xai_responses",
        model_id="grok-4-1-fast-reasoning",
        notes="xAI Responses API; fast reasoning model",
    ),
    "llama-3.1-8b-instant": ModelSpec(
        key="llama-3.1-8b-instant",
        display="Llama 3.1 8B Instant 128k",
        provider="groq_chat",
        model_id="llama-3.1-8b-instant",
        notes="Groq Chat Completions",
    ),
}

MIGRATED_DEFAULT_SELECTION = {
    "instructors": ["llama-3.1-8b-instant"],
    "targets": ["gpt-5.4-nano"],
    "prompt_type": "1",
    "questionnaire_mode": "120-180",
}


# =============================================================================
# Logging and caching
# =============================================================================

def setup_logging(output_dir: Path) -> logging.Logger:
    output_dir.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("multi_model_validation")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    formatter = logging.Formatter("%(asctime)s  %(message)s")
    file_handler = logging.FileHandler(str(output_dir / "multi_model_validation.log"), encoding="utf-8")
    file_handler.setFormatter(formatter)
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    return logger


def stable_hash(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def slugify(text: str) -> str:
    return re.sub(r"[^a-zA-Z0-9_.-]+", "_", text).strip("_").lower()


def model_short_name(model_key: str) -> str:
    """Short readable names for Excel sheet names. Excel sheet names are limited to 31 chars."""
    aliases = {
        "gpt-5.4": "gpt54",
        "gpt-5.4-mini": "gpt54mini",
        "gpt-5.4-nano": "gpt54nano",
        "gpt-realtime-1.5": "realtime15",
        "gemma-4-26b-a4b-it": "gemma26b",
        "gemma-4-31b-it": "gemma31b",
        "grok-4.20-reasoning": "grok420",
        "grok-4-1-fast-reasoning": "grok41fast",
        "llama-3.1-8b-instant": "llama31_8b",
    }
    return aliases.get(model_key, slugify(model_key)[:12])


def excel_safe_sheet_name(raw_name: str, used_names: Optional[set[str]] = None) -> str:
    """Make an Excel-safe sheet name, preserving as much model identity as possible."""
    invalid = "[]:*?/\\"
    cleaned = "".join("_" if ch in invalid else ch for ch in raw_name).strip() or "Sheet"
    cleaned = cleaned[:31]
    if used_names is None:
        return cleaned
    candidate = cleaned
    counter = 2
    while candidate in used_names:
        suffix = f"_{counter}"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        counter += 1
    used_names.add(candidate)
    return candidate


def pair_sheet_name(instructor_model: str, target_model: str, used_names: Optional[set[str]] = None) -> str:
    return excel_safe_sheet_name(
        f"{model_short_name(instructor_model)}_to_{model_short_name(target_model)}",
        used_names=used_names,
    )



def combo_short_label(selection: Dict[str, List[str]]) -> str:
    """Build a compact folder label from the selected instructor/target model combination(s)."""
    instructors = selection.get("instructors", [])
    targets = selection.get("targets", [])
    prompt_suffix = f"p{selection.get('prompt_type', '1')}_{selection.get('questionnaire_mode', '120-180').replace('-', 'to')}"
    pair_labels = [
        f"{model_short_name(instr)}_to_{model_short_name(targ)}"
        for instr in instructors
        for targ in targets
    ]
    if len(pair_labels) == 1:
        return f"{pair_labels[0]}_{prompt_suffix}"
    if len(pair_labels) <= 3:
        return "__".join(pair_labels) + f"_{prompt_suffix}"
    return f"multi_{len(pair_labels)}_combos_{prompt_suffix}"


def next_trial_number(output_root: Path = OUTPUT_ROOT) -> int:
    """Return the next sequential trial number based on existing trial folders."""
    output_root.mkdir(parents=True, exist_ok=True)
    max_seen = 0
    pattern = re.compile(r"^trial_(\d+)_")
    for child in output_root.iterdir():
        if not child.is_dir():
            continue
        match = pattern.match(child.name)
        if match:
            try:
                max_seen = max(max_seen, int(match.group(1)))
            except ValueError:
                pass
    return max_seen + 1


def make_trial_output_dir(selection: Dict[str, List[str]], timestamp: str) -> Path:
    """Create a folder name that contains the trial number and selected model combo."""
    trial = next_trial_number(OUTPUT_ROOT)
    combo = slugify(combo_short_label(selection))[:80] or "model_combo"
    return OUTPUT_ROOT / f"trial_{trial:03d}_{combo}_{timestamp}"


FACET_SHORT_NAMES = {
    "Anxiety": "Anx",
    "Friendliness": "Frnd",
    "Imagination": "Imag",
    "Trust": "Trst",
    "Self-Efficacy": "SEff",
    "Anger": "Ang",
    "Gregariousness": "Greg",
    "Artistic Interests": "Art",
    "Morality": "Mor",
    "Orderliness": "Ord",
    "Depression": "Dep",
    "Assertiveness": "Asrt",
    "Emotionality": "Emo",
    "Altruism": "Alt",
    "Dutifulness": "Duty",
    "Self-Consciousness": "SelfC",
    "Activity Level": "Act",
    "Adventurousness": "Adv",
    "Cooperation": "Coop",
    "Achievement-Striving": "Ach",
    "Immoderation": "Immod",
    "Excitement-Seeking": "Exc",
    "Intellect": "Int",
    "Modesty": "Mod",
    "Self-Discipline": "SDisc",
    "Vulnerability": "Vuln",
    "Cheerfulness": "Cheer",
    "Liberalism": "Lib",
    "Sympathy": "Sym",
    "Cautiousness": "Caut",
}


def facet_short_name(facet: str) -> str:
    return FACET_SHORT_NAMES.get(facet, re.sub(r"[^A-Za-z0-9]+", "", facet)[:8] or facet[:8])

def facet_alignment_col(facet: str) -> str:
    return "aligned_facet_" + re.sub(r"[^a-zA-Z0-9]+", "_", facet).strip("_")


def load_json_cache(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_json_cache(path: Path, data: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)


INSTRUCTOR_CACHE_FILE = CACHE_DIR / "instructor_prompt_cache.json"
SYSTEM_PROMPT_CACHE_FILE = CACHE_DIR / "validation_system_prompt_cache.json"
TARGET_BATCH_CACHE_FILE = CACHE_DIR / "target_batch_scores_cache.json"
RAW_TEXT_CACHE_FILE = CACHE_DIR / "raw_text_generation_cache.json"

INSTRUCTOR_CACHE = load_json_cache(INSTRUCTOR_CACHE_FILE)
SYSTEM_PROMPT_CACHE = load_json_cache(SYSTEM_PROMPT_CACHE_FILE)
TARGET_BATCH_CACHE = load_json_cache(TARGET_BATCH_CACHE_FILE)
RAW_TEXT_CACHE = load_json_cache(RAW_TEXT_CACHE_FILE)


# =============================================================================
# Selection menu
# =============================================================================

def load_previous_selection() -> Dict[str, Any]:
    if SELECTION_FILE.exists():
        try:
            data = json.loads(SELECTION_FILE.read_text(encoding="utf-8"))
            instructors = [m for m in data.get("instructors", []) if m in AVAILABLE_MODELS]
            targets = [m for m in data.get("targets", []) if m in AVAILABLE_MODELS]
            prompt_type = str(data.get("prompt_type", "1")).strip()
            if prompt_type not in {"1", "2", "3"}:
                prompt_type = "1"
            questionnaire_mode = str(data.get("questionnaire_mode", "120-180")).strip()
            if questionnaire_mode not in {"120-180", "180-120"}:
                questionnaire_mode = "120-180"
            if instructors and targets:
                return {"instructors": instructors, "targets": targets, "prompt_type": prompt_type, "questionnaire_mode": questionnaire_mode}
        except Exception:
            pass
    return MIGRATED_DEFAULT_SELECTION.copy()


def save_selection(selection: Dict[str, Any]) -> None:
    SELECTION_FILE.write_text(json.dumps(selection, indent=2), encoding="utf-8")


def print_model_menu() -> None:
    print("\nAvailable models for BOTH instructor and target roles:")
    for idx, key in enumerate(AVAILABLE_MODELS.keys(), start=1):
        spec = AVAILABLE_MODELS[key]
        print(f"  {idx}. {spec.key:<24} | {spec.display:<24} | {spec.notes}")


def format_model_list(model_keys: Sequence[str]) -> str:
    return ", ".join(model_keys)


def parse_model_selection(raw: str) -> List[str]:
    raw = raw.strip()
    keys = list(AVAILABLE_MODELS.keys())
    if raw.lower() in {"all", "*"}:
        return keys
    selected: List[str] = []
    for part in raw.split(","):
        token = part.strip()
        if not token:
            continue
        if token.isdigit():
            idx = int(token)
            if not (1 <= idx <= len(keys)):
                raise ValueError(f"Model number {idx} is out of range.")
            selected.append(keys[idx - 1])
        else:
            token = token.lower()
            if token not in AVAILABLE_MODELS:
                raise ValueError(f"Unknown model: {token}")
            selected.append(token)
    deduped = []
    for m in selected:
        if m not in deduped:
            deduped.append(m)
    if not deduped:
        raise ValueError("No valid models selected.")
    return deduped


def prompt_type_description(prompt_type: str) -> str:
    pt = str(prompt_type)
    if pt == "2":
        return "Prompt Type 2 — FPS prompt path: instructor and target prompts from the uploaded FPS validation script, adapted to chosen split"
    if pt == "3":
        return "Prompt Type 3 — FPS instructor + original/current target prompt"
    return "Prompt Type 1 — current validation system: direct second-person validation prompt"


def choose_prompt_type_interactively(selection: Dict[str, Any]) -> Dict[str, Any]:
    current = str(selection.get("prompt_type", "1")).strip()
    if current not in {"1", "2", "3"}:
        current = "1"

    print("\n" + "=" * 86)
    print("PROMPT TYPE SELECTION — still no participant data is sent before this step")
    print("=" * 86)
    print("Current prompt type:", current)
    print(" ", prompt_type_description(current))
    print("\nChoose prompt system:")
    print("  1. Prompt Type 1 — current system")
    print("     Direct second-person validation prompt generated specifically for the target LLM.")
    print("  2. Prompt Type 2 — FPS prompt path")
    print("     Uses the same FPS-style instructor narrative prompt and target adoption prompt; for 180-120, item-range wording is adapted.")
    print("  3. Prompt Type 3 — FPS instructor + original target prompt")
    print("     Instructor uses the uploaded FPS-style prompt, but target uses the original/current validation target prompt.")

    raw = input("Prompt type [1/2/3, Enter keeps current]: ").strip()
    if raw not in {"1", "2", "3"}:
        raw = current
    selection["prompt_type"] = raw

    save_selection(selection)
    print("Using", prompt_type_description(raw))
    return choose_questionnaire_mode_interactively(selection)


def choose_questionnaire_mode_interactively(selection: Dict[str, Any]) -> Dict[str, Any]:
    current = str(selection.get("questionnaire_mode", "120-180")).strip()
    if current not in {"120-180", "180-120"}:
        current = "120-180"

    print("\n" + "=" * 86)
    print("QUESTIONNAIRE SPLIT SELECTION — still no participant data is sent before this step")
    print("=" * 86)
    print("Current questionnaire mode:", current)
    print(" ", questionnaire_mode_description(current))
    print("\nChoose questionnaire split:")
    print("  1. 120-180 — target sees labelled items 1-120 and answers items 121-300")
    print("  2. 180-120 — target sees labelled items 1-180 and answers items 181-300")

    raw = input("Questionnaire mode [1/2, Enter keeps current]: ").strip()
    if raw == "2":
        mode = "180-120"
    elif raw == "1":
        mode = "120-180"
    else:
        mode = current

    selection["questionnaire_mode"] = mode
    save_selection(selection)
    print("Using", questionnaire_mode_description(mode))
    return selection


def choose_models_interactively() -> Dict[str, Any]:
    selection = load_previous_selection()

    print("\n" + "=" * 86)
    print("MODEL SELECTION — no participant data is sent before this step")
    print("=" * 86)
    print("Current chosen instructor model(s):", format_model_list(selection["instructors"]))
    print("Current chosen target model(s)    :", format_model_list(selection["targets"]))
    print("Current prompt type               :", selection.get("prompt_type", "1"), "-", prompt_type_description(str(selection.get("prompt_type", "1"))))
    print("Current questionnaire mode        :", selection.get("questionnaire_mode", "120-180"), "-", questionnaire_mode_description(str(selection.get("questionnaire_mode", "120-180"))))
    print("\nPress 1 to continue with the same model choice")
    print("Press 2 to change model choice")
    choice = input("Choice: ").strip() or "1"

    if choice == "1":
        print("Using previous/current model choice.")
        return choose_prompt_type_interactively(selection)

    if choice != "2":
        print("Unrecognised choice. Using previous/current model choice.")
        return choose_prompt_type_interactively(selection)

    print_model_menu()
    print("\nEnter comma-separated numbers or model IDs. Type 'all' to run every model.")
    while True:
        try:
            instructor_raw = input("Instructor model(s): ").strip()
            instructors = parse_model_selection(instructor_raw)
            target_raw = input("Target model(s): ").strip()
            targets = parse_model_selection(target_raw)
            new_selection = {"instructors": instructors, "targets": targets, "prompt_type": str(selection.get("prompt_type", "1")), "questionnaire_mode": str(selection.get("questionnaire_mode", "120-180"))}
            save_selection(new_selection)
            print("\nSaved model choice.")
            print("Instructor model(s):", format_model_list(instructors))
            print("Target model(s)    :", format_model_list(targets))
            return choose_prompt_type_interactively(new_selection)
        except ValueError as e:
            print(f"Selection error: {e}")
            print("Try again.")

def ask_max_participants(default: int) -> Optional[int]:
    raw = input(f"How many participants to validate? Press Enter for {default}, or type 'all': ").strip()
    if not raw:
        return default
    if raw.lower() == "all":
        return None
    try:
        value = int(raw)
        if value <= 0:
            raise ValueError
        return value
    except ValueError:
        print(f"Invalid participant count. Using {default}.")
        return default


# =============================================================================
# Environment validation and lazy clients
# =============================================================================

def selected_providers(selection: Dict[str, List[str]]) -> set:
    providers = set()
    for key in selection["instructors"] + selection["targets"]:
        providers.add(AVAILABLE_MODELS[key].provider)
    return providers


def validate_environment(selection: Dict[str, List[str]]) -> None:
    providers = selected_providers(selection)
    missing = []
    if "openai_responses" in providers or "openai_realtime" in providers:
        if not OPENAI_API_KEY:
            missing.append("OPENAI_API_KEY")
    if "xai_responses" in providers:
        if not XAI_API_KEY:
            missing.append("XAI_API_KEY")
    if "groq_chat" in providers:
        if not GROQ_API_KEY:
            missing.append("GROQ_API_KEY or GROQ_API_KEY_2 or GROQ_API_KEY_3 or GROQ_API_KEY_TUD")
    if "google_genai" in providers:
        if not GEMINI_API_KEY:
            missing.append("GEMINI_API_KEY or GOOGLE_AI_STUDIO_API_KEY or GOOGLE_API_KEY")
    if missing:
        raise RuntimeError("Missing required environment variable(s): " + ", ".join(missing))

    if "google_genai" in providers:
        print("\nGoogle AI Studio / Gemini API selected for Gemma.")
        print(f"  GEMINI_API_KEY / GOOGLE_AI_STUDIO_API_KEY / GOOGLE_API_KEY = {'set' if GEMINI_API_KEY else 'not set'}\n")
    if "xai_responses" in providers:
        print("\nxAI / Grok Responses API selected.")
        print(f"  XAI_API_KEY = {'set' if XAI_API_KEY else 'not set'}")
        print(f"  XAI_BASE_URL = {XAI_BASE_URL}\n")


_OPENAI_CLIENT = None
_XAI_CLIENT = None
_ASYNC_OPENAI_CLIENT = None
_GROQ_CLIENT = None
_GROQ_CLIENT_KEY_NAME: Optional[str] = None
_GROQ_KEY_INDEX = 0
_GOOGLE_GENAI_CLIENT = None

_GROQ_THROTTLE_LOCK = threading.Lock()
_GROQ_CALL_HISTORY: List[Tuple[float, int, str]] = []
_GROQ_LAST_CALL_AT = 0.0


def get_openai_client():
    global _OPENAI_CLIENT
    if _OPENAI_CLIENT is None:
        from openai import OpenAI
        _OPENAI_CLIENT = OpenAI(api_key=OPENAI_API_KEY, max_retries=0)
    return _OPENAI_CLIENT


def get_xai_client():
    global _XAI_CLIENT
    if _XAI_CLIENT is None:
        from openai import OpenAI
        _XAI_CLIENT = OpenAI(api_key=XAI_API_KEY, base_url=XAI_BASE_URL, max_retries=0)
    return _XAI_CLIENT


def get_async_openai_client():
    global _ASYNC_OPENAI_CLIENT
    if _ASYNC_OPENAI_CLIENT is None:
        from openai import AsyncOpenAI
        _ASYNC_OPENAI_CLIENT = AsyncOpenAI(api_key=OPENAI_API_KEY, max_retries=0)
    return _ASYNC_OPENAI_CLIENT


def get_current_groq_key_entry() -> Tuple[str, str]:
    if not GROQ_API_KEY_ENTRIES:
        return "GROQ_API_KEY", ""
    idx = max(0, min(_GROQ_KEY_INDEX, len(GROQ_API_KEY_ENTRIES) - 1))
    return GROQ_API_KEY_ENTRIES[idx]


def get_groq_client():
    global _GROQ_CLIENT, _GROQ_CLIENT_KEY_NAME
    key_name, api_key = get_current_groq_key_entry()
    if _GROQ_CLIENT is None or _GROQ_CLIENT_KEY_NAME != key_name:
        from groq import Groq
        _GROQ_CLIENT = Groq(api_key=api_key, max_retries=0)
        _GROQ_CLIENT_KEY_NAME = key_name
    return _GROQ_CLIENT


def current_groq_key_name() -> str:
    return get_current_groq_key_entry()[0]


def rotate_groq_key(logger: logging.Logger, reason: str = "") -> bool:
    """Switch to the next configured Groq key at runtime.

    Used for long retry windows that usually indicate an exhausted daily/token
    budget on the current key. It clears local TPM history because the next key
    may have a separate quota window. If keys belong to the same organization,
    Groq may still enforce shared org limits, but this gives the requested fallback.
    """
    global _GROQ_KEY_INDEX, _GROQ_CLIENT, _GROQ_CLIENT_KEY_NAME, _GROQ_CALL_HISTORY, _GROQ_LAST_CALL_AT

    if len(GROQ_API_KEY_ENTRIES) <= 1:
        return False

    old_name = current_groq_key_name()
    for step in range(1, len(GROQ_API_KEY_ENTRIES) + 1):
        candidate_index = (_GROQ_KEY_INDEX + step) % len(GROQ_API_KEY_ENTRIES)
        candidate_name, _ = GROQ_API_KEY_ENTRIES[candidate_index]
        if candidate_name != old_name:
            _GROQ_KEY_INDEX = candidate_index
            _GROQ_CLIENT = None
            _GROQ_CLIENT_KEY_NAME = None
            with _GROQ_THROTTLE_LOCK:
                _GROQ_CALL_HISTORY = []
                _GROQ_LAST_CALL_AT = 0.0
            logger.warning(
                f"    Groq key switch: {old_name} -> {candidate_name}"
                + (f" | reason: {reason}" if reason else "")
            )
            return True
    return False


def get_google_genai_client():
    global _GOOGLE_GENAI_CLIENT
    if _GOOGLE_GENAI_CLIENT is None:
        from google import genai
        _GOOGLE_GENAI_CLIENT = genai.Client(api_key=GEMINI_API_KEY)
    return _GOOGLE_GENAI_CLIENT


# =============================================================================
# Retry and parsing helpers
# =============================================================================

def is_rate_limit_error(exc: Exception) -> bool:
    text = str(exc).lower()
    return ("429" in text) or ("rate limit" in text) or ("too many requests" in text)


def is_temporary_server_error(exc: Exception) -> bool:
    text = str(exc).lower()
    return any(phrase in text for phrase in [
        "500", "502", "503", "504",
        "internal server error", "bad gateway", "service unavailable",
        "gateway timeout", "connection reset", "timed out", "timeout",
        "connection error", "server disconnected",
    ])


def is_realtime_empty_output_error(exc: Exception) -> bool:
    """OpenAI Realtime can occasionally complete without a text item.
    Treat it as a transient provider failure and retry the same target batch.
    """
    text = str(exc).lower()
    return (
        "realtime response completed without text output" in text
        or "realtime connection closed before response.done" in text
        or "completed without text output" in text
    )


def safe_get_header(exc: Exception, key: str) -> Optional[str]:
    try:
        response = getattr(exc, "response", None)
        headers = getattr(response, "headers", None)
        if headers:
            return headers.get(key)
    except Exception:
        return None
    return None


def _parse_retry_after_seconds_from_text(text: str) -> Optional[float]:
    """Extract Groq-style retry hints such as 'Please try again in 3.5s'."""
    lowered = (text or "").lower()
    m = re.search(r"try again in\s+([0-9]+(?:\.[0-9]+)?)\s*s", lowered)
    if m:
        return float(m.group(1))
    m = re.search(r"try again in\s+([0-9]+(?:\.[0-9]+)?)\s*m\s*([0-9]+(?:\.[0-9]+)?)?\s*s?", lowered)
    if m:
        minutes = float(m.group(1))
        seconds = float(m.group(2) or 0)
        return minutes * 60 + seconds
    return None


def compute_retry_delay(exc: Exception, attempt: int, default_base: float = 3.0, default_cap: float = 120.0) -> float:
    retry_after = safe_get_header(exc, "retry-after")
    if retry_after:
        try:
            return max(1.0, float(retry_after))
        except ValueError:
            pass

    parsed = _parse_retry_after_seconds_from_text(str(exc))
    if parsed is not None:
        # Add a small cushion so we do not hit the same TPM edge again.
        return max(1.0, parsed + 1.0)

    exponential = min(default_base * (2 ** attempt), default_cap)
    return exponential + random.uniform(0, 1.5)


def retry_hint_seconds(exc: Exception) -> Optional[float]:
    retry_after = safe_get_header(exc, "retry-after")
    if retry_after:
        try:
            return float(retry_after)
        except ValueError:
            pass
    return _parse_retry_after_seconds_from_text(str(exc))


def is_groq_single_request_too_large(exc: Exception) -> bool:
    """Detect Groq's 413 TPM/request-size error for one oversized request."""
    text = str(exc).lower()
    return (
        ("413" in text and "request too large" in text)
        or ("requested" in text and "limit" in text and "tokens per minute" in text and "tpm" in text)
    )


def should_rotate_groq_key(exc: Exception) -> bool:
    if not is_rate_limit_error(exc) and not is_groq_single_request_too_large(exc):
        return False
    if is_groq_single_request_too_large(exc):
        return len(GROQ_API_KEY_ENTRIES) > 1
    hint = retry_hint_seconds(exc)
    if hint is None:
        return False
    return hint >= GROQ_KEY_ROTATE_AFTER_SECONDS


def estimate_groq_requested_tokens(system_prompt: str, user_prompt: str, max_tokens: int) -> int:
    # This is an intentionally conservative estimate. Groq's error message reports
    # requested tokens as input tokens + max output allowance.
    char_count = len(system_prompt or "") + len(user_prompt or "")
    estimated_input = int(math.ceil(char_count / max(GROQ_EST_CHARS_PER_TOKEN, 1.0)))
    return estimated_input + int(max_tokens or 0) + GROQ_TOKEN_BUFFER


def _prune_groq_history(now: float) -> None:
    global _GROQ_CALL_HISTORY
    _GROQ_CALL_HISTORY = [(ts, toks, purpose) for ts, toks, purpose in _GROQ_CALL_HISTORY if now - ts < 60.0]


def reserve_groq_tpm_capacity(
    system_prompt: str,
    user_prompt: str,
    max_tokens: int,
    logger: logging.Logger,
    purpose: str,
) -> int:
    """Wait before Groq calls so llama-3.1-8b-instant stays under TPM limits.

    This is proactive throttling. It prevents the repeated 429 pattern where the call is
    small enough for context but too large for the current 60-second TPM window.
    """
    global _GROQ_LAST_CALL_AT

    estimate = estimate_groq_requested_tokens(system_prompt, user_prompt, max_tokens)
    if not GROQ_THROTTLE_ENABLED:
        return estimate

    hard_limit = max(1, GROQ_TPM_LIMIT)
    safe_limit = max(1, int(hard_limit * GROQ_TPM_SAFETY))

    # If one request is itself above the current key/account TPM limit, waiting cannot make
    # that single request legal on this key. Do not add it to the rolling TPM history and do
    # not sleep for 60s. Let call_groq_chat try it immediately; on Groq's 413/TPM error,
    # call_model_text will rotate to the next configured Groq key.
    effective_limit = max(safe_limit, min(estimate, hard_limit))
    if estimate > hard_limit:
        logger.warning(
            f"    Groq request estimate for {purpose} is {estimate} tokens, above GROQ_TPM_LIMIT={hard_limit}. "
            "Skipping proactive sleep and trying configured Groq keys immediately. "
            "If every key has the same TPM limit, this exact Prompt Type 2 Llama request cannot run on Groq on-demand; "
            "use a higher-limit Groq key, a non-Groq instructor, or Prompt Type 1/3."
        )
        return estimate

    while True:
        with _GROQ_THROTTLE_LOCK:
            now = time.time()
            _prune_groq_history(now)
            used = sum(toks for _, toks, _ in _GROQ_CALL_HISTORY)
            wait_for_interval = max(0.0, GROQ_MIN_INTERVAL_SECONDS - (now - _GROQ_LAST_CALL_AT))

            wait_for_tokens = 0.0
            if used + estimate > effective_limit and _GROQ_CALL_HISTORY:
                oldest_ts = min(ts for ts, _, _ in _GROQ_CALL_HISTORY)
                wait_for_tokens = max(0.0, 60.0 - (now - oldest_ts) + 0.75)

            wait_seconds = max(wait_for_interval, wait_for_tokens)
            if wait_seconds <= 0.0:
                _GROQ_CALL_HISTORY.append((now, estimate, purpose))
                _GROQ_LAST_CALL_AT = now
                return estimate

        logger.info(
            f"    Groq throttle before {purpose}: waiting {wait_seconds:.1f}s "
            f"(estimated_request={estimate}, recent_used={used}, safe_limit={effective_limit})"
        )
        time.sleep(wait_seconds)


def extract_json_candidate(text: str) -> str:
    cleaned = (text or "").strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```$", "", cleaned)
    if cleaned.startswith("{") and cleaned.endswith("}"):
        return cleaned
    if cleaned.startswith("[") and cleaned.endswith("]"):
        return cleaned

    obj_start, obj_end = cleaned.find("{"), cleaned.rfind("}")
    arr_start, arr_end = cleaned.find("["), cleaned.rfind("]")
    candidates = []
    if obj_start != -1 and obj_end > obj_start:
        candidates.append(cleaned[obj_start:obj_end + 1])
    if arr_start != -1 and arr_end > arr_start:
        candidates.append(cleaned[arr_start:arr_end + 1])
    if not candidates:
        raise ValueError(f"No JSON object or array found in output: {cleaned[:250]}")
    # Prefer object because OpenAI structured output returns {"scores": [...]}.
    candidates.sort(key=lambda c: 0 if c.startswith("{") else 1)
    return candidates[0]


def parse_scores_json(raw_text: str, expected_n: int) -> List[int]:
    """Parse target-model questionnaire scores robustly.

    Preferred output is either {"scores": [...]} or a raw JSON array.
    Some models, especially Llama/Groq under Prompt Type 2, may instead
    return a JSON object keyed by item number, e.g. {"181": 5, "182": 4}.
    This parser accepts that form and sorts numeric keys in ascending order.
    """
    snippet = extract_json_candidate(raw_text)
    parsed = json.loads(snippet)

    if isinstance(parsed, dict):
        scores = parsed.get("scores")

        # Fallback: accept {"181": 5, "182": 4, ...} or {"1": 5, "2": 4, ...}.
        if scores is None:
            numeric_items = []
            for key, value in parsed.items():
                key_text = str(key).strip()
                if key_text.isdigit():
                    numeric_items.append((int(key_text), value))

            if len(numeric_items) == expected_n and len(numeric_items) == len(parsed):
                numeric_items.sort(key=lambda kv: kv[0])
                scores = [value for _, value in numeric_items]
    elif isinstance(parsed, list):
        scores = parsed
    else:
        raise ValueError("JSON output must be a scores object, item-keyed object, or array.")

    if not isinstance(scores, list):
        raise ValueError(f"JSON output missing scores list or item-number keyed scores: {str(parsed)[:250]}")
    if len(scores) != expected_n:
        raise ValueError(f"Expected {expected_n} scores, got {len(scores)}.")

    clean_scores: List[int] = []
    for value in scores:
        if isinstance(value, str) and value.strip().isdigit():
            value = int(value.strip())
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        if not isinstance(value, int) or value < 1 or value > 5:
            raise ValueError(f"Invalid score value: {value} in {scores}")
        clean_scores.append(value)
    return clean_scores


def extract_response_text(response: Any) -> str:
    direct = getattr(response, "output_text", None)
    if direct:
        return direct.strip()

    chunks = []
    for item in getattr(response, "output", []) or []:
        for content in getattr(item, "content", []) or []:
            ctype = getattr(content, "type", None)
            if ctype in {"output_text", "text"}:
                piece = getattr(content, "text", None)
                if piece:
                    chunks.append(piece)
    return "".join(chunks).strip()


def build_scores_schema(n: int) -> Dict[str, Any]:
    return {
        "type": "json_schema",
        "name": f"personality_scores_{n}",
        "strict": True,
        "schema": {
            "type": "object",
            "properties": {
                "scores": {
                    "type": "array",
                    "items": {"type": "integer", "minimum": 1, "maximum": 5},
                    "minItems": n,
                    "maxItems": n,
                }
            },
            "required": ["scores"],
            "additionalProperties": False,
        },
    }


def log_openai_prompt_cache_usage(response: Any, logger: logging.Logger, label: str) -> None:
    try:
        usage = getattr(response, "usage", None)
        details = getattr(usage, "input_tokens_details", None) if usage else None
        input_tokens = getattr(usage, "input_tokens", None) if usage else None
        cached_tokens = getattr(details, "cached_tokens", None) if details else None
        if input_tokens is not None and cached_tokens is not None:
            logger.info(f"    OpenAI prompt cache {label}: cached_tokens={cached_tokens} / input_tokens={input_tokens}")
    except Exception:
        pass


# =============================================================================
# Model call layer
# =============================================================================

def call_model_text(
    spec: ModelSpec,
    system_prompt: str,
    user_prompt: str,
    logger: logging.Logger,
    purpose: str,
    max_tokens: int,
    expected_scores: Optional[int] = None,
) -> str:
    cache_key = stable_hash(
        json.dumps({
            "prompt_version": PROMPT_VERSION,
            "model_key": spec.key,
            "provider": spec.provider,
            "purpose": purpose,
            "expected_scores": expected_scores,
            "system_prompt": system_prompt,
            "user_prompt": user_prompt,
        }, ensure_ascii=False, sort_keys=True)
    )
    # Cache raw free-text generations such as instructor prompts.
    # For score batches, cache only after successful JSON parsing in TARGET_BATCH_CACHE;
    # otherwise a malformed raw response could poison future runs.
    if expected_scores is None:
        cached = RAW_TEXT_CACHE.get(cache_key)
        if cached is not None:
            logger.info(f"    raw text cache hit: {purpose} | {spec.key}")
            return cached

    groq_oversize_tried_keys: set[str] = set()

    for attempt in range(8):
        try:
            if spec.provider == "openai_responses":
                text = call_openai_responses(spec, system_prompt, user_prompt, logger, purpose, max_tokens, expected_scores)
            elif spec.provider == "xai_responses":
                text = call_xai_responses(spec, system_prompt, user_prompt, logger, purpose, max_tokens, expected_scores)
            elif spec.provider == "openai_realtime":
                text = asyncio.run(call_openai_realtime(spec, system_prompt, user_prompt, logger, purpose, max_tokens))
            elif spec.provider == "groq_chat":
                text = call_groq_chat(spec, system_prompt, user_prompt, logger, purpose, max_tokens, expected_scores)
            elif spec.provider == "google_genai":
                text = call_google_genai(spec, system_prompt, user_prompt, logger, purpose, max_tokens, expected_scores)
            else:
                raise RuntimeError(f"Unsupported provider: {spec.provider}")

            if expected_scores is None:
                RAW_TEXT_CACHE[cache_key] = text
                save_json_cache(RAW_TEXT_CACHE_FILE, RAW_TEXT_CACHE)
            return text

        except Exception as e:
            if spec.provider == "groq_chat" and is_groq_single_request_too_large(e):
                failed_key = current_groq_key_name()
                groq_oversize_tried_keys.add(failed_key)
                if len(groq_oversize_tried_keys) < len(GROQ_API_KEY_ENTRIES):
                    switched = rotate_groq_key(
                        logger,
                        reason=f"single Groq request too large on {failed_key} during {purpose}",
                    )
                    if switched:
                        logger.warning(
                            f"    retry {attempt + 1} for {purpose} | {spec.key}: {failed_key} rejected the single request; "
                            f"trying {current_groq_key_name()} immediately without 60s sleep"
                        )
                        time.sleep(0.25)
                        continue
                raise RuntimeError(
                    f"Groq rejected this single request as larger than the TPM/request limit for all tried keys "
                    f"({', '.join(sorted(groq_oversize_tried_keys))}). Waiting will not fix this. "
                    f"For exact Prompt Type 2 with Llama as instructor, use a Groq key/org with a higher TPM limit, "
                    f"lower GROQ_TYPE2_INSTRUCTOR_MAX_OUTPUT_TOKENS, switch to Prompt Type 1/3, or use GPT/Gemini as instructor. "
                    f"Original error: {repr(e)}"
                ) from e

            if spec.provider == "groq_chat" and should_rotate_groq_key(e):
                hint = retry_hint_seconds(e)
                switched = rotate_groq_key(
                    logger,
                    reason=f"long Groq retry hint {hint:.1f}s during {purpose}" if hint is not None else f"Groq rate limit during {purpose}",
                )
                if switched:
                    if hint is not None:
                        logger.warning(
                            f"    retry {attempt + 1} for {purpose} | {spec.key}: switched to {current_groq_key_name()} instead of waiting {hint:.1f}s"
                        )
                    else:
                        logger.warning(
                            f"    retry {attempt + 1} for {purpose} | {spec.key}: switched to {current_groq_key_name()}"
                        )
                    time.sleep(1.0)
                    continue

            if is_realtime_empty_output_error(e):
                # This is usually a transient Realtime API completion issue, not a bad score.
                # Retry the same target batch; do not fail the whole case immediately.
                delay = min(2.0 + attempt * 1.5, 10.0) + random.uniform(0, 0.5)
                logger.warning(
                    f"    retry {attempt + 1} for {purpose} | {spec.key} after empty Realtime output | waiting {delay:.1f}s"
                )
                time.sleep(delay)
                continue

            if is_rate_limit_error(e) or is_temporary_server_error(e):
                delay = compute_retry_delay(e, attempt)
                logger.warning(f"    retry {attempt + 1} for {purpose} | {spec.key} after error: {repr(e)} | waiting {delay:.1f}s")
                time.sleep(delay)
                continue
            raise

    raise RuntimeError(f"Failed after repeated retries: {purpose} | {spec.key}")


def call_openai_responses(
    spec: ModelSpec,
    system_prompt: str,
    user_prompt: str,
    logger: logging.Logger,
    purpose: str,
    max_tokens: int,
    expected_scores: Optional[int],
) -> str:
    client = get_openai_client()
    input_items = []
    if system_prompt and system_prompt.strip():
        input_items.append({
            "type": "message",
            "role": "system",
            "content": [{"type": "input_text", "text": system_prompt}],
        })
    input_items.append({
        "type": "message",
        "role": "user",
        "content": [{"type": "input_text", "text": user_prompt}],
    })

    kwargs: Dict[str, Any] = {
        "model": spec.model_id,
        "temperature": 0,
        "max_output_tokens": max_tokens,
        "prompt_cache_key": stable_hash(system_prompt),
        "prompt_cache_retention": "24h",
        "input": input_items,
    }
    if DEFAULT_REASONING_EFFORT.lower() not in {"", "none-disabled", "disabled"}:
        kwargs["reasoning"] = {"effort": DEFAULT_REASONING_EFFORT}
    if expected_scores is not None:
        kwargs["text"] = {"format": build_scores_schema(expected_scores)}

    response = client.responses.create(**kwargs)
    log_openai_prompt_cache_usage(response, logger, purpose)
    text = extract_response_text(response)
    if not text:
        raise RuntimeError("OpenAI Responses API returned empty text.")
    return text



def call_xai_responses(
    spec: ModelSpec,
    system_prompt: str,
    user_prompt: str,
    logger: logging.Logger,
    purpose: str,
    max_tokens: int,
    expected_scores: Optional[int],
) -> str:
    """Call xAI/Grok through the OpenAI-compatible Responses API.

    xAI documents OpenAI REST API compatibility at https://api.x.ai/v1. We use the
    OpenAI Python client with base_url=XAI_BASE_URL and API key XAI_API_KEY.
    """
    client = get_xai_client()
    input_items = []
    if system_prompt and system_prompt.strip():
        input_items.append({
            "type": "message",
            "role": "system",
            "content": [{"type": "input_text", "text": system_prompt}],
        })
    input_items.append({
        "type": "message",
        "role": "user",
        "content": [{"type": "input_text", "text": user_prompt}],
    })

    kwargs: Dict[str, Any] = {
        "model": spec.model_id,
        "temperature": 0,
        "max_output_tokens": max_tokens,
        "input": input_items,
    }

    # Try structured outputs for score batches. If xAI rejects the schema or
    # the selected Grok model does not support it, retry once without schema;
    # the existing local parser will still validate the returned scores.
    if expected_scores is not None:
        kwargs["text"] = {"format": build_scores_schema(expected_scores)}

    try:
        response = client.responses.create(**kwargs)
    except Exception as e:
        if expected_scores is not None and any(
            token in str(e).lower()
            for token in ["schema", "response_format", "text.format", "unsupported", "invalid"]
        ):
            logger.warning(
                f"    xAI structured-output fallback for {purpose} | {spec.key}: {repr(e)}"
            )
            kwargs.pop("text", None)
            fallback_prompt = (
                f"{user_prompt}\n\n"
                "OUTPUT FORMAT OVERRIDE:\n"
                f"Return exactly {expected_scores} scores only. Use plain comma-separated integers only, e.g. 4,2,5,3.\n"
                "Do not include item numbers, item text, labels, explanations, markdown, JSON, brackets, quotes, comments, or extra values."
            )
            kwargs["input"] = []
            if system_prompt and system_prompt.strip():
                kwargs["input"].append({
                    "type": "message",
                    "role": "system",
                    "content": [{"type": "input_text", "text": system_prompt}],
                })
            kwargs["input"].append({
                "type": "message",
                "role": "user",
                "content": [{"type": "input_text", "text": fallback_prompt}],
            })
            response = client.responses.create(**kwargs)
        else:
            raise

    text = extract_response_text(response)
    if not text:
        raise RuntimeError("xAI Responses API returned empty text.")
    return text
async def call_openai_realtime(
    spec: ModelSpec,
    system_prompt: str,
    user_prompt: str,
    logger: logging.Logger,
    purpose: str,
    max_tokens: int,
) -> str:
    client = get_async_openai_client()
    async with client.realtime.connect(model=spec.model_id) as connection:
        await connection.session.update(
            session={
                "type": "realtime",
                "output_modalities": ["text"],
            }
        )

        await connection.response.create(
            response={
                "conversation": "none",
                "output_modalities": ["text"],
                "max_output_tokens": max_tokens,
                "input": ([{
                    "type": "message",
                    "role": "system",
                    "content": [{"type": "input_text", "text": system_prompt}],
                }] if system_prompt and system_prompt.strip() else []) + [{
                    "type": "message",
                    "role": "user",
                    "content": [{"type": "input_text", "text": user_prompt}],
                }],
            }
        )

        deltas: List[str] = []
        final_text: Optional[str] = None
        async for event in connection:
            etype = getattr(event, "type", None)
            if etype == "response.output_text.delta":
                deltas.append(getattr(event, "delta", "") or "")
            elif etype == "response.output_text.done":
                text_value = getattr(event, "text", None)
                if text_value:
                    final_text = text_value
            elif etype == "error":
                err = getattr(event, "error", None)
                message = getattr(err, "message", "Unknown Realtime API error")
                code = getattr(err, "code", None)
                raise RuntimeError(f"Realtime API error{f' [{code}]' if code else ''}: {message}")
            elif etype == "response.done":
                if final_text:
                    return final_text.strip()
                combined = "".join(deltas).strip()
                if combined:
                    return combined
                # last-resort event dump parser
                try:
                    payload = event.model_dump() if hasattr(event, "model_dump") else {}
                    chunks = []
                    for item in (payload.get("response") or {}).get("output") or []:
                        for content in item.get("content") or []:
                            if content.get("type") in {"output_text", "text"} and content.get("text"):
                                chunks.append(content["text"])
                    fallback = "".join(chunks).strip()
                    if fallback:
                        return fallback
                except Exception:
                    pass
                raise RuntimeError("Realtime response completed without text output.")
    raise RuntimeError("Realtime connection closed before response.done.")


def call_groq_chat(
    spec: ModelSpec,
    system_prompt: str,
    user_prompt: str,
    logger: logging.Logger,
    purpose: str,
    max_tokens: int,
    expected_scores: Optional[int],
) -> str:
    client = get_groq_client()
    messages = []
    if system_prompt and system_prompt.strip():
        messages.append({"role": "system", "content": system_prompt})
    messages.append({"role": "user", "content": user_prompt})
    kwargs: Dict[str, Any] = {
        "model": spec.model_id,
        "messages": messages,
        "temperature": 0,
        "max_tokens": max_tokens,
    }
    if expected_scores is not None:
        kwargs["response_format"] = {"type": "json_object"}

    estimated_tokens = reserve_groq_tpm_capacity(
        system_prompt=system_prompt,
        user_prompt=user_prompt,
        max_tokens=max_tokens,
        logger=logger,
        purpose=purpose,
    )
    logger.info(f"    Groq call {purpose} | key={current_groq_key_name()} | estimated_request_tokens={estimated_tokens}")

    try:
        response = client.chat.completions.create(**kwargs)
    except Exception as e:
        # Some OpenAI-compatible and Groq account configs may reject response_format.
        if expected_scores is not None and "response_format" in str(e).lower():
            kwargs.pop("response_format", None)
            response = client.chat.completions.create(**kwargs)
        else:
            raise

    text = response.choices[0].message.content
    if not text:
        raise RuntimeError("Groq returned empty text.")
    return text.strip()


def call_google_genai(
    spec: ModelSpec,
    system_prompt: str,
    user_prompt: str,
    logger: logging.Logger,
    purpose: str,
    max_tokens: int,
    expected_scores: Optional[int],
) -> str:
    """Call Gemma through Google AI Studio / Gemini API using the google-genai SDK."""
    client = get_google_genai_client()
    from google.genai import types

    if expected_scores is not None:
        prompt = (
            f"{user_prompt}\n\n"
            f"Return only valid JSON in this exact shape: "
            f"{{\"scores\": [integer, integer, ...]}} with exactly {expected_scores} integers from 1 to 5."
        )
        config = types.GenerateContentConfig(
            temperature=0,
            max_output_tokens=max_tokens,
            system_instruction=(system_prompt if system_prompt and system_prompt.strip() else None),
            response_mime_type="application/json",
            response_schema={
                "type": "object",
                "properties": {
                    "scores": {
                        "type": "array",
                        "items": {"type": "integer", "minimum": 1, "maximum": 5},
                        "minItems": expected_scores,
                        "maxItems": expected_scores,
                    }
                },
                "required": ["scores"],
                "additionalProperties": False,
            },
        )
    else:
        prompt = user_prompt
        config = types.GenerateContentConfig(
            temperature=0,
            max_output_tokens=max_tokens,
            system_instruction=(system_prompt if system_prompt and system_prompt.strip() else None),
        )

    try:
        response = client.models.generate_content(
            model=spec.model_id,
            contents=prompt,
            config=config,
        )
    except Exception as e:
        # Some Google model/server combinations may reject response_schema.
        # Retry with JSON MIME type only and let our parser validate the result.
        if expected_scores is not None and "schema" in str(e).lower():
            fallback_config = types.GenerateContentConfig(
                temperature=0,
                max_output_tokens=max_tokens,
                system_instruction=(system_prompt if system_prompt and system_prompt.strip() else None),
                response_mime_type="application/json",
            )
            response = client.models.generate_content(
                model=spec.model_id,
                contents=prompt,
                config=fallback_config,
            )
        else:
            raise

    text = getattr(response, "text", None)
    if not text:
        # Last-resort parser for SDK versions that expose candidates/parts.
        try:
            chunks = []
            for cand in getattr(response, "candidates", []) or []:
                content = getattr(cand, "content", None)
                for part in getattr(content, "parts", []) or []:
                    piece = getattr(part, "text", None)
                    if piece:
                        chunks.append(piece)
            text = "".join(chunks).strip()
        except Exception:
            text = ""
    if not text:
        raise RuntimeError("Google AI Studio / Gemini API returned empty text.")
    return text.strip()


# =============================================================================
# Scoring and prompt building
# =============================================================================

def normalize_raw_score(value: Any) -> int:
    try:
        raw = int(value)
    except Exception:
        raw = 3
    return raw if 1 <= raw <= 5 else 3


def score_item(raw_score: Any, reverse: bool) -> int:
    raw = normalize_raw_score(raw_score)
    return 6 - raw if reverse else raw


def facet_level_label(value: float) -> str:
    if value < 1.5:
        return "very low"
    if value < 2.5:
        return "low"
    if value < 3.5:
        return "mixed / moderate"
    if value < 4.5:
        return "high"
    return "very high"


def get_scores(row: pd.Series, items: Sequence[int]) -> List[int]:
    return [normalize_raw_score(row.get(f"i{item}", 3)) for item in items]


def summarize_scale_distribution(scores: Sequence[int]) -> Dict[str, Any]:
    clean = [normalize_raw_score(s) for s in scores]
    counts = {score: clean.count(score) for score in range(1, 6)}
    total = len(clean) if clean else 1
    return {
        "mean": round(float(np.mean(clean)), 4),
        "variance": round(float(np.var(clean)), 4),
        "neutral_pct": round(counts[3] / total * 100, 2),
        "extreme_pct": round((counts[1] + counts[5]) / total * 100, 2),
        "low_pct": round((counts[1] + counts[2]) / total * 100, 2),
        "high_pct": round((counts[4] + counts[5]) / total * 100, 2),
        "counts": counts,
    }


def compute_domain_avgs_from_row(row: pd.Series, key_entries: Sequence[Tuple[int, str, str, bool, str]]) -> Dict[str, float]:
    totals = {domain: 0 for domain in DOMAIN_ORDER}
    counts = {domain: 0 for domain in DOMAIN_ORDER}
    for item_num, domain, facet, reverse, item_text in key_entries:
        totals[domain] += score_item(row.get(f"i{item_num}", 3), reverse)
        counts[domain] += 1
    return {domain: round(totals[domain] / counts[domain], 4) for domain in DOMAIN_ORDER}


def compute_domain_avgs_from_scores(scores: Sequence[int], key_entries: Sequence[Tuple[int, str, str, bool, str]], base_item: int) -> Dict[str, float]:
    totals = {domain: 0 for domain in DOMAIN_ORDER}
    counts = {domain: 0 for domain in DOMAIN_ORDER}
    for item_num, domain, facet, reverse, item_text in key_entries:
        idx = item_num - base_item
        totals[domain] += score_item(scores[idx], reverse)
        counts[domain] += 1
    return {domain: round(totals[domain] / counts[domain], 4) for domain in DOMAIN_ORDER}


def compute_facet_avgs_from_scores(scores: Sequence[int], key_entries: Sequence[Tuple[int, str, str, bool, str]], base_item: int) -> Dict[str, Optional[float]]:
    """Facet mean scores for a train/test split after applying the official reverse-scoring key."""
    totals = {facet: 0 for facet in FACET_ORDER}
    counts = {facet: 0 for facet in FACET_ORDER}
    for item_num, domain, facet, reverse, item_text in key_entries:
        idx = item_num - base_item
        if 0 <= idx < len(scores):
            totals[facet] += score_item(scores[idx], reverse)
            counts[facet] += 1
    return {
        facet: round(totals[facet] / counts[facet], 4) if counts[facet] else None
        for facet in FACET_ORDER
    }


def score_sequence_by_items(scores: Sequence[int], items: Sequence[int]) -> List[int]:
    """Return trait-direction scores for the given items, applying reverse coding where needed.

    The target LLM still answers raw Likert values. Validation comparisons use this
    reverse-coded version for both human and model scores, so alignment scores operate on the personality-direction scale.
    """
    scored: List[int] = []
    for idx, item_num in enumerate(items):
        scored.append(score_item(scores[idx], ITEM_TO_REVERSE[item_num]))
    return scored


def compute_facet_avgs_from_row(row: pd.Series) -> Dict[str, float]:
    totals = {facet: 0 for facet in FACET_ORDER}
    counts = {facet: 0 for facet in FACET_ORDER}
    for item_num, domain, facet, reverse, item_text in ITEM_KEY_300:
        totals[facet] += score_item(row.get(f"i{item_num}", 3), reverse)
        counts[facet] += 1
    return {facet: round(totals[facet] / counts[facet], 4) for facet in FACET_ORDER}


def build_profile_summary(row: pd.Series) -> Dict[str, Any]:
    full_scores = get_scores(row, ALL_ITEMS)
    dist = summarize_scale_distribution(full_scores)
    ocean_avgs = compute_domain_avgs_from_row(row, ITEM_KEY_300)
    facet_avgs = compute_facet_avgs_from_row(row)

    ordered_facets = sorted(facet_avgs.items(), key=lambda kv: (kv[1], kv[0]))
    lowest = ordered_facets[:6]
    highest = list(reversed(ordered_facets[-6:]))

    distribution_lines = [f"- score {score}: {dist['counts'][score]} of 300" for score in range(1, 6)]
    ocean_lines = [f"- {DOMAIN_FULL_NAMES[d]} ({d}): avg={ocean_avgs[d]}" for d in DOMAIN_ORDER]
    facet_lines = [
        f"- {facet} ({FACET_TO_DOMAIN[facet]}): avg={facet_avgs[facet]} ({facet_level_label(facet_avgs[facet])})"
        for facet in FACET_ORDER
    ]
    highest_lines = [f"- {facet}: {score}" for facet, score in highest]
    lowest_lines = [f"- {facet}: {score}" for facet, score in lowest]

    summary_text = "\n".join([
        "PROFILE SUMMARY FROM ITEMS 1-300",
        f"- overall mean raw response: {dist['mean']}",
        f"- raw response variance: {dist['variance']}",
        f"- neutral (3) usage: {dist['neutral_pct']}%",
        f"- extreme (1 or 5) usage: {dist['extreme_pct']}%",
        f"- low-end raw response (1 or 2) usage: {dist['low_pct']}%",
        f"- high-end raw response (4 or 5) usage: {dist['high_pct']}%",
        "- raw score distribution:",
        *distribution_lines,
        "- OCEAN averages from all 300 items after reverse scoring where required:",
        *ocean_lines,
        "- highest scoring facets after reverse scoring:",
        *highest_lines,
        "- lowest scoring facets after reverse scoring:",
        *lowest_lines,
        "- all facet averages after reverse scoring:",
        *facet_lines,
    ])

    return {
        "summary_text": summary_text,
        "ocean_avgs": ocean_avgs,
        "facet_avgs": facet_avgs,
        "highest_facets": highest,
        "lowest_facets": lowest,
    }


def build_instructor_evidence(row: pd.Series) -> str:
    """Detailed item-level evidence. Useful for high-token instructor models only."""
    profile = build_profile_summary(row)
    facet_blocks: List[str] = []

    for facet in FACET_ORDER:
        domain = FACET_TO_DOMAIN[facet]
        facet_avg = profile["facet_avgs"][facet]
        lines = [
            f"Facet: {facet} | Domain: {domain} | reverse-scored facet average: {facet_avg} ({facet_level_label(facet_avg)})"
        ]
        for item_num, item_domain, item_facet, reverse, item_text in ITEM_KEY_300:
            if item_facet != facet:
                continue
            raw = normalize_raw_score(row.get(f"i{item_num}", 3))
            interpreted = score_item(raw, reverse)
            reverse_note = "reverse-scored item" if reverse else "direct-scored item"
            lines.append(
                f"  - i{item_num}. {item_text} -> raw response: {LABEL_MAP_FORMAL[raw]}; "
                f"interpreted tendency score: {interpreted}/5 ({reverse_note})"
            )
        facet_blocks.append("\n".join(lines))

    return profile["summary_text"] + "\n\nDETAILED FACET EVIDENCE FROM ALL 300 ITEMS\n\n" + "\n\n".join(facet_blocks)


def build_compact_instructor_evidence(row: pd.Series) -> str:
    """Compact evidence for rate-limited instructor models such as Groq on-demand Llama.

    This uses all 300 questionnaire responses after official reverse scoring, but it sends
    only aggregate OCEAN/facet summaries rather than all item-level evidence. This keeps
    the request under small TPM limits while preserving the validation signal needed to
    generate a detailed facet-by-facet control prompt.
    """
    profile = build_profile_summary(row)
    dist = summarize_scale_distribution(get_scores(row, ALL_ITEMS))
    ocean_avgs = profile["ocean_avgs"]
    facet_avgs = profile["facet_avgs"]
    highest = profile["highest_facets"]
    lowest = profile["lowest_facets"]

    ocean_lines = [
        f"- {DOMAIN_FULL_NAMES[d]} ({d}): {ocean_avgs[d]:.3f} / 5"
        for d in DOMAIN_ORDER
    ]
    facet_lines = [
        f"- {facet} ({FACET_TO_DOMAIN[facet]}): {facet_avgs[facet]:.3f} / 5 = {facet_level_label(facet_avgs[facet])}"
        for facet in FACET_ORDER
    ]
    highest_lines = [f"- {facet}: {score:.3f}" for facet, score in highest]
    lowest_lines = [f"- {facet}: {score:.3f}" for facet, score in lowest]
    distribution_lines = [f"- raw score {score}: {dist['counts'][score]} of 300" for score in range(1, 6)]

    return "\n".join([
        "COMPACT FULL-PROFILE EVIDENCE FROM ALL 300 IPIP-NEO RESPONSES",
        "The values below are computed from the official item key with reverse scoring applied where required.",
        f"Overall raw-response mean: {dist['mean']} | variance: {dist['variance']}",
        f"Neutral usage: {dist['neutral_pct']}% | extreme usage: {dist['extreme_pct']}% | low-end usage: {dist['low_pct']}% | high-end usage: {dist['high_pct']}%",
        "Raw response distribution:",
        *distribution_lines,
        "OCEAN domain averages:",
        *ocean_lines,
        "Highest facets:",
        *highest_lines,
        "Lowest facets:",
        *lowest_lines,
        "All 30 facet averages:",
        *facet_lines,
    ])


def build_instructor_user_prompt_type1(row: pd.Series, compact: bool = False) -> str:
    evidence = build_compact_instructor_evidence(row) if compact else build_instructor_evidence(row)
    compact_note = (
        "The evidence is compact aggregate evidence, not item-level evidence. "
        "Still produce a detailed facet-by-facet target prompt from the OCEAN and facet profile.\n\n"
        if compact else ""
    )
    return (
        "Convert the questionnaire evidence below into a SECOND-PERSON personality control prompt for a target LLM.\n\n"
        f"{compact_note}"
        "The final prompt you write will be given directly to the target LLM. Therefore, write as if you are speaking to the target LLM itself.\n"
        "Use direct second-person wording throughout: 'you are this person', 'you take this personality', 'you enjoy...', 'you tend to...', 'you behave...', 'you avoid...', 'you respond...'.\n\n"
        "Required structure for your output:\n"
        "1. Start with a clear adoption line similar to: 'For this validation task, you are this person and you take on this personality as your own.'\n"
        "2. Write a concise natural overview in second person: how you think, speak, feel, interact, decide, cooperate, resist, and regulate yourself.\n"
        "3. Include exactly one behavioural line for each of the 30 facets, each written as a direct 'you' instruction.\n"
        "4. Include response-style calibration: how strongly or moderately you use 1, 2, 3, 4, and 5-style answers.\n"
        "5. Include guardrails that prevent exaggeration, especially around warmth, trust, artistry, emotional depth, ambition, caution, and cooperation.\n"
        "6. Do not mention IPIP, Big Five, OCEAN, questionnaire, dataset, item numbers, scoring, labels, or hidden answers in the final prompt.\n"
        "7. Preserve intensity carefully. Do not convert moderate tendencies into extreme ones.\n"
        "8. Make the output directly usable as a system/developer prompt for a target model.\n"
        "9. End with a short testing instruction in second person: when you take the next questionnaire items, answer based on the personality you have now, not by predicting hidden data.\n"
        "10. Keep the output concise enough for repeated validation calls: aim for 900-1300 words.\n\n"
        "Evidence:\n"
        f"{evidence}"
    )


def build_instructor_user_prompt_type2(row: pd.Series, compact: bool = False) -> str:
    """Exact FPS-method instructor prompt from the uploaded/pasted FPS method script.

    Prompt Type 2 intentionally mirrors the original FPS script's instructor prompt:
    all 300 raw questionnaire responses are grouped by facet, translated to the
    same intensity labels, and passed into the same second-person narrative request.

    This function intentionally ignores the compact flag. With a 6000 TPM Groq
    on-demand limit, Llama may reject this exact prompt because the request itself
    can exceed the minute token limit.
    """
    all_scores = [normalize_raw_score(row.get(f"i{n}", 3)) for n in range(1, 301)]

    facet_data: Dict[str, List[Tuple[str, int]]] = {}
    for item_num, raw_score in zip(range(1, 301), all_scores):
        facet = ITEM_TO_FACET[item_num]
        item_text = ITEM_TO_TEXT[item_num]
        facet_data.setdefault(facet, []).append((item_text, raw_score))

    score_label = {
        1: "almost never / strongly disagree",
        2: "rarely / slightly disagree",
        3: "sometimes / neutral",
        4: "often / slightly agree",
        5: "almost always / strongly agree",
    }

    ocean_sections = {
        'N': 'Emotional Tendencies',
        'E': 'Social and Energetic Style',
        'O': 'Curiosity and Openness to Experience',
        'A': 'Interpersonal Style and Values',
        'C': 'Work Style and Self-Regulation',
    }

    facet_block = ""
    for ocean_dim, section_title in ocean_sections.items():
        facet_block += f"\n=== {section_title} ===\n"
        for facet in FACET_ORDER:
            dim = FACET_TO_DOMAIN[facet]
            if dim != ocean_dim:
                continue
            if facet not in facet_data:
                continue
            items_in_facet = facet_data[facet]
            avg_score = sum(s for _, s in items_in_facet) / len(items_in_facet)
            avg_label = score_label[round(avg_score)]
            sorted_items = sorted(items_in_facet, key=lambda x: x[1])
            facet_block += f"\n[{facet}] — overall level: {avg_label}\n"
            for item_text, score in sorted_items:
                facet_block += f'  - "{item_text}" → {score_label.get(score, "neutral")}\n'

    prompt = (
        f"Based on the following questionnaire responses, write a detailed "
        f"second-person personality description."
        f"For each behaviour described,"
        f"be very specific about the INTENSITY and FREQUENCY — use words like "
        f"'always', 'almost never', 'strongly', 'slightly', 'rarely', "
        f"'consistently', 'occasionally' to preserve exactly how strongly "
        f"each trait and facet applies. For example, do not say 'I am social' — instead "
        f"say 'I almost always feel energised around people and very strongly "
        f"prefer being in groups over being alone'. The intensity words must "
        f"map clearly to these levels: "
        f"'almost never / strongly disagree' = level 1, "
        f"'rarely / slightly disagree' = level 2, "
        f"'sometimes / neutral' = level 3, "
        f"'often / slightly agree' = level 4, "
        f"'almost always / strongly agree' = level 5. "
        f"Do NOT mention any personality frameworks, trait names, OCEAN, "
        f"Big Five, scores, or psychological terminology, but mention facets if necessary"
        f"Write as natural flowing prose."
        f"{facet_block}"
    )
    return prompt


def build_instructor_user_prompt(row: pd.Series, compact: bool = False, prompt_type: str = "1") -> str:
    # Prompt Type 2 and Type 3 both use the uploaded FPS-style instructor prompt.
    # Type 2 pairs it with the FPS-style target prompt.
    # Type 3 pairs it with the original/current validation target prompt.
    if str(prompt_type) in {"2", "3"}:
        return build_instructor_user_prompt_type2(row, compact=compact)
    return build_instructor_user_prompt_type1(row, compact=compact)


def build_instructor_system_prompt(prompt_type: str = "1") -> str:
    if str(prompt_type) == "2":
        # The uploaded FPS script sends the instructor prompt as a user-only prompt.
        # Empty system prompt allows provider wrappers to omit the system message.
        return ""
    if str(prompt_type) == "3":
        return (
            "You are a careful personality narrative generator for validation experiments. "
            "Follow the FPS validation style: write a detailed second-person personality description, "
            "preserve intensity and frequency precisely, and avoid mentioning tests, scores, Big Five, OCEAN, hidden labels, or datasets. "
            "Do not exaggerate. Do not invent biography, job, life history, preferences, trauma, culture, or demographic details. "
            "Keep nearby concepts separate: imagination is not artistic devotion; intellect is not emotional depth; "
            "trust is not sympathy; dutifulness is not ambition; cooperation is not friendliness."
        )
    return (
        "You are a careful personality-prompt generator for validation experiments. "
        "Your job is to transform structured questionnaire evidence into a faithful, restrained, "
        "facet-specific personality control prompt written directly to a target LLM in second person. "
        "The target prompt must tell the model: you are this person, you take this personality, "
        "you behave this way, and you answer later questionnaire items from inside this adopted personality. "
        "Preserve nuance. Do not exaggerate. Do not invent biography, job, life history, preferences, trauma, culture, or demographic details. "
        "Keep nearby concepts separate: imagination is not artistic devotion; intellect is not emotional depth; "
        "trust is not sympathy; dutifulness is not ambition; cooperation is not friendliness."
    )


def generate_instructor_prompt(row: pd.Series, instructor_key: str, logger: logging.Logger, prompt_type: str = "1") -> str:
    spec = AVAILABLE_MODELS[instructor_key]
    # Groq on-demand accounts can have small TPM limits even for 128k-context models.
    # Use compact aggregate evidence for Groq/Llama so the instructor call does not exceed TPM.
    # Other providers may use detailed item-level evidence unless FORCE_COMPACT_INSTRUCTOR=1.
    force_compact = (os.getenv("FORCE_COMPACT_INSTRUCTOR", "0").strip() == "1")
    compact = force_compact or (spec.provider == "groq_chat")
    user_prompt = build_instructor_user_prompt(row, compact=compact, prompt_type=prompt_type)
    system_prompt = build_instructor_system_prompt(prompt_type=prompt_type)
    case_id = str(row.get("case", "unknown"))

    cache_key = stable_hash(json.dumps({
        "prompt_version": PROMPT_VERSION,
        "prompt_type": str(prompt_type),
        "role": "instructor",
        "case": case_id,
        "model_key": instructor_key,
        "system_prompt": system_prompt,
        "user_prompt": user_prompt,
    }, ensure_ascii=False, sort_keys=True))

    cached = INSTRUCTOR_CACHE.get(cache_key)
    if cached is not None:
        logger.info(f"    instructor cache hit | case {case_id} | {instructor_key} | prompt_type={prompt_type}")
        return cached

    instructor_max_tokens = DEFAULT_INSTRUCTOR_MAX_TOKENS
    if str(prompt_type) == "2" and spec.provider == "groq_chat":
        # The pasted FPS script used max_tokens=400 for the Llama narrative.
        # Using the normal 1800-token validation default makes the requested TPM much larger.
        instructor_max_tokens = GROQ_TYPE2_INSTRUCTOR_MAX_OUTPUT_TOKENS

    logger.info(
        f"    generating instructor prompt | case {case_id} | {instructor_key} | prompt_type={prompt_type} | "
        f"compact={compact} | chars={len(system_prompt) + len(user_prompt)} | max_tokens={instructor_max_tokens}"
    )
    text = call_model_text(
        spec=spec,
        system_prompt=system_prompt,
        user_prompt=user_prompt,
        logger=logger,
        purpose=f"instructor_prompt_case_{case_id}_ptype_{prompt_type}",
        max_tokens=instructor_max_tokens,
        expected_scores=None,
    )
    INSTRUCTOR_CACHE[cache_key] = text
    save_json_cache(INSTRUCTOR_CACHE_FILE, INSTRUCTOR_CACHE)
    return text

def build_training_calibration(row: pd.Series) -> str:
    train_scores = get_scores(row, TRAIN_ITEMS)
    dist = summarize_scale_distribution(train_scores)
    trait_avgs = compute_domain_avgs_from_row(row, TRAIN_KEY)

    facet_totals = {facet: [] for facet in FACET_ORDER}
    for item_num, domain, facet, reverse, item_text in TRAIN_KEY:
        facet_totals[facet].append(score_item(row.get(f"i{item_num}", 3), reverse))

    facet_lines = []
    for facet in FACET_ORDER:
        vals = facet_totals.get(facet, [])
        if vals:
            avg = round(float(np.mean(vals)), 3)
            facet_lines.append(f"- {facet}: avg={avg} ({facet_level_label(avg)})")
    distribution_lines = [f"- score {score}: {dist['counts'][score]} of {len(TRAIN_ITEMS)}" for score in range(1, 6)]
    trait_lines = [f"- {DOMAIN_FULL_NAMES[d]} ({d}): avg={trait_avgs[d]}" for d in DOMAIN_ORDER]

    return "\n".join([
        f"RESPONSE STYLE CALIBRATION FROM {TRAIN_RANGE_LABEL.upper()}",
        f"- overall mean raw response: {dist['mean']}",
        f"- raw response variance: {dist['variance']}",
        f"- neutral (3) usage: {dist['neutral_pct']}%",
        f"- extreme (1 or 5) usage: {dist['extreme_pct']}%",
        f"- low-end raw response (1 or 2) usage: {dist['low_pct']}%",
        f"- high-end raw response (4 or 5) usage: {dist['high_pct']}%",
        "- raw score distribution:",
        *distribution_lines,
        "- training-domain averages from the selected training items after reverse scoring:",
        *trait_lines,
        "- training-facet averages from the selected training items after reverse scoring:",
        *facet_lines,
    ])


def build_few_shot_train(row: pd.Series, prompt_type: str = "1") -> str:
    examples = []
    for item_num, domain, facet, reverse, item_text in TRAIN_KEY:
        raw = normalize_raw_score(row.get(f"i{item_num}", 3))
        label = LABEL_MAP_FORMAL.get(raw, "Neither Accurate Nor Inaccurate")
        examples.append(f"{item_num}. {item_text} -> {label}")

    if str(prompt_type) == "2" and QUESTIONNAIRE_MODE == "120-180":
        header = """For this request only, adopt one internally consistent personality.
    The narrative below describes your temporary personality for this case.
    The 120 labeled examples below show how this same adopted person typically responds to questionnaire statements.
    Use them as anchors for your own response style.

    You are not analysing a dataset and you are not trying to predict a hidden respondent's labels.
    You are answering as the adopted person.

    Do not invent biography or backstory.
    Do not mention personality theories, factor names, score values, or test mechanics.
    When answering later items, stay behaviourally consistent with both the narrative description and the labeled examples.

    Known questionnaire items and labels for this adopted person (items 1-120 only):
    """
        return header + "\n".join(examples)

    if str(prompt_type) == "2":
        # Same FPS few-shot structure, adapted only for the optional 180-120 split.
        header = f"""For this request only, adopt one internally consistent personality.
    The narrative below describes your temporary personality for this case.
    The {len(TRAIN_ITEMS)} labeled examples below show how this same adopted person typically responds to questionnaire statements.
    Use them as anchors for your own response style.

    You are not analysing a dataset and you are not trying to predict a hidden respondent's labels.
    You are answering as the adopted person.

    Do not invent biography or backstory.
    Do not mention personality theories, factor names, score values, or test mechanics.
    When answering later items, stay behaviourally consistent with both the narrative description and the labeled examples.

    Known questionnaire items and labels for this adopted person ({TRAIN_RANGE_LABEL} only):
    """
        return header + "\n".join(examples)

    return (
        f"Your known questionnaire responses for this adopted personality ({TRAIN_RANGE_LABEL} only). "
        "These examples show how you have responded and the response style you should continue with:\n"
        + "\n".join(examples)
    )


def build_validation_system_prompt(
    instructor_prompt: str,
    calibration_prompt: str,
    few_shot_prompt: str,
    instructor_key: str,
    target_key: str,
    case_id: str,
    logger: logging.Logger,
    prompt_type: str = "1",
) -> str:
    cache_key = stable_hash(json.dumps({
        "prompt_version": PROMPT_VERSION,
        "prompt_type": str(prompt_type),
        "questionnaire_mode": QUESTIONNAIRE_MODE,
        "instructor_key": instructor_key,
        "target_key": target_key,
        "case_id": case_id,
        "instructor_prompt": instructor_prompt,
        "calibration_prompt": calibration_prompt,
        "few_shot_prompt": few_shot_prompt,
    }, ensure_ascii=False, sort_keys=True))

    cached = SYSTEM_PROMPT_CACHE.get(cache_key)
    if cached is not None:
        logger.info(f"    system prompt cache hit | case {case_id} | {instructor_key} -> {target_key} | prompt_type={prompt_type}")
        return cached

    if str(prompt_type) == "2":
        prompt = (
            "For this request only, temporarily adopt the personality, behavioural tendencies, "
            "emotional style, interpersonal style, and response habits described below as if they are your own. "
            "You are not analysing a respondent, predicting someone else's answers, or reconstructing hidden labels. "
            "You are to answer as the person whose personality is represented by the information below.\n\n"

            "Treat the narrative profile as a high-level description of your own enduring personality. "
            f"Treat the labeled examples from {TRAIN_RANGE_LABEL} as concrete evidence of how you, in this temporary adopted personality, "
            "typically respond to questionnaire statements.\n\n"

            "When answering later items:\n"
            "- answer in first-person psychological continuity, as if this personality is your current one\n"
            "- do NOT try to infer or guess what the original human respondent selected\n"
            "- do NOT optimise for matching a hidden answer key\n"
            "- do NOT mention the training items, test split, questionnaire design, personality frameworks, or scores\n"
            "- preserve the demonstrated response style, including intensity, moderation, and consistency\n"
            "- if the narrative and labeled examples differ, prefer the labeled examples for item-level response tendencies while staying globally consistent with the narrative\n"
            "- this adopted personality applies ONLY to the current case and current request\n\n"

            "--- PERSONALITY NARRATIVE FROM THE 300-ITEM PROFILE ---\n"
            f"{instructor_prompt}\n\n"
            f"--- LABELED REFERENCE ITEMS ({TRAIN_RANGE_LABEL}) FOR THIS SAME ADOPTED PERSONA ---\n"
            f"{few_shot_prompt}"
        )
    else:
        # Prompt Type 1 and Prompt Type 3 use the original/current validation target prompt.
        # For Prompt Type 3, the instructor prompt itself is FPS-style, but the target prompt
        # below remains this original/current target wrapper.
        prompt = (
            "For this validation request only, you are this person. You take this personality as your own for the current case. "
            "You think, speak, feel, choose, cooperate, resist, and answer as this adopted person would.\n\n"

            "You are not analysing a respondent, predicting someone else's hidden labels, reconstructing an answer key, "
            "or optimising for a dataset. You are answering from inside the personality you have now.\n\n"

            "Core validation rules:\n"
            "- You answer in first-person psychological continuity, as if this personality is your current one.\n"
            "- You stay consistent with the instructor-generated second-person personality prompt.\n"
            "- You use the labelled training items as concrete anchors for your item-level response tendencies.\n"
            "- You preserve the demonstrated response style, including intensity, moderation, neutrality, and extremity.\n"
            "- If the high-level personality prompt and labelled examples conflict, you prefer the labelled examples for item-level style while preserving global personality coherence.\n"
            "- You do not mention the training items, test split, questionnaire design, IPIP, Big Five, OCEAN, scores, hidden data, or labels.\n"
            "- You do not invent biography or backstory.\n"
            "- This adopted personality applies only to the current validation case.\n\n"

            "--- SECOND-PERSON PERSONALITY CONTROL PROMPT FROM ITEMS 1-300 ---\n"
            f"{instructor_prompt}\n\n"

            "--- RESPONSE STYLE CALIBRATION FROM THE SELECTED TRAINING ITEMS ---\n"
            f"{calibration_prompt}\n\n"

            "--- YOUR LABELLED REFERENCE TRAINING RESPONSES ---\n"
            f"{few_shot_prompt}"
        )

    SYSTEM_PROMPT_CACHE[cache_key] = prompt
    save_json_cache(SYSTEM_PROMPT_CACHE_FILE, SYSTEM_PROMPT_CACHE)
    return prompt

def build_target_batch_instruction(item_numbers: Sequence[int], prompt_type: str = "1") -> str:
    numbered = "\n".join(f"{item_num}. {ITEM_TO_TEXT[item_num]}" for item_num in item_numbers)
    n = len(item_numbers)
    if str(prompt_type) == "2":
        # Prompt Type 2 uses the FPS-style target batch instruction with a raw JSON array.
        # Prompt Type 3 intentionally falls through to the original/current target instruction.
        return (
            "You are now taking the following questionnaire items while embodying the adopted personality described in the system message.\n"
            "Answer these items as that person would genuinely answer them now.\n"
            "Do NOT predict a hidden respondent's answer key. Do NOT analyse the dataset. Simply respond from within the adopted personality.\n\n"
            "Use this scale strictly:\n"
            "1 = Very Inaccurate\n"
            "2 = Moderately Inaccurate\n"
            "3 = Neither Accurate Nor Inaccurate\n"
            "4 = Moderately Accurate\n"
            "5 = Very Accurate\n\n"
            f"Return ONLY a JSON array of EXACTLY {n} integers, one per statement in order. "
            "No explanation, no prose, no markdown.\n\n"
            f"{numbered}"
        )

    return (
        "You are now taking the following questionnaire items based on the personality you have now. "
        "You are this person for the current validation case, and you answer as this person would genuinely answer.\n"
        "Do NOT predict a hidden respondent's answer key. Do NOT analyse the dataset. "
        "Do NOT try to match expected labels. Simply answer from within the personality you have adopted.\n\n"
        "Use this scale strictly:\n"
        "1 = Very Inaccurate\n"
        "2 = Moderately Inaccurate\n"
        "3 = Neither Accurate Nor Inaccurate\n"
        "4 = Moderately Accurate\n"
        "5 = Very Accurate\n\n"
        f"Return ONLY valid JSON in this exact shape: {{\"scores\": [/* exactly {n} integers */]}}.\n"
        f"The scores list must contain EXACTLY {n} integers, one per statement, in order.\n"
        "No explanation, no prose, no markdown.\n\n"
        f"{numbered}"
    )


def run_target_batches(
    row: pd.Series,
    system_prompt: str,
    target_key: str,
    instructor_key: str,
    logger: logging.Logger,
    batch_size: int,
    prompt_type: str = "1",
) -> List[int]:
    spec = AVAILABLE_MODELS[target_key]
    case_id = str(row.get("case", "unknown"))
    all_scores: List[int] = []

    effective_batch_size = batch_size
    if spec.provider == "groq_chat":
        effective_batch_size = max(1, min(batch_size, LLAMA_TARGET_BATCH_SIZE))
        if effective_batch_size != batch_size:
            logger.info(
                f"    Llama/Groq target batch override: using {effective_batch_size} items per call "
                f"instead of {batch_size}. Set LLAMA_TARGET_BATCH_SIZE in .env to change this."
            )

    for start in range(0, len(TEST_ITEMS), effective_batch_size):
        item_numbers = TEST_ITEMS[start:start + effective_batch_size]
        batch_label = f"case_{case_id}_items_{item_numbers[0]}_{item_numbers[-1]}"
        instruction = build_target_batch_instruction(item_numbers, prompt_type=prompt_type)
        expected_n = len(item_numbers)

        cache_key = stable_hash(json.dumps({
            "prompt_version": PROMPT_VERSION,
            "questionnaire_mode": QUESTIONNAIRE_MODE,
            "prompt_type": str(prompt_type),
            "role": "target_batch",
            "case_id": case_id,
            "instructor_key": instructor_key,
            "target_key": target_key,
            "batch_items": item_numbers,
            "system_prompt": system_prompt,
            "instruction": instruction,
        }, ensure_ascii=False, sort_keys=True))

        cached = TARGET_BATCH_CACHE.get(cache_key)
        if cached is not None:
            logger.info(f"    target batch cache hit | {batch_label} | {target_key}")
            all_scores.extend(cached)
            continue

        logger.info(f"    target answering {batch_label} | {target_key}")
        raw = call_model_text(
            spec=spec,
            system_prompt=system_prompt,
            user_prompt=instruction,
            logger=logger,
            purpose=f"target_scores_{batch_label}",
            max_tokens=(
                DEFAULT_OPENAI_MAX_OUTPUT_TOKENS
                if spec.provider.startswith("openai")
                else min(DEFAULT_GROQ_MAX_TOKENS, max(160, expected_n * 8))
                if spec.provider == "groq_chat"
                else DEFAULT_GEMINI_MAX_OUTPUT_TOKENS
                if spec.provider == "google_genai"
                else DEFAULT_OPENAI_MAX_OUTPUT_TOKENS
            ),
            expected_scores=expected_n,
        )
        scores = parse_scores_json(raw, expected_n)
        TARGET_BATCH_CACHE[cache_key] = scores
        save_json_cache(TARGET_BATCH_CACHE_FILE, TARGET_BATCH_CACHE)
        all_scores.extend(scores)

        time.sleep(0.25)

    if len(all_scores) != len(TEST_ITEMS):
        raise RuntimeError(f"Target returned {len(all_scores)} total test scores, expected {len(TEST_ITEMS)}.")
    return all_scores


# =============================================================================
# Metrics
# =============================================================================

def safe_pearson(x: Sequence[int], y: Sequence[int]) -> Tuple[Optional[float], Optional[float]]:
    try:
        if len(set(x)) < 2 or len(set(y)) < 2:
            return None, None
        r, p = pearsonr(np.array(x, dtype=float), np.array(y, dtype=float))
        return round(float(r), 4), round(float(p), 6)
    except Exception:
        return None, None


def compute_scored_alignment_by_domain(human_scored_test: Sequence[int], model_scored_test: Sequence[int]) -> Dict[str, float]:
    """MPI-style aligned score by OCEAN domain after reverse coding both sides.

    Each difference is computed on scored values, not raw response values.
    For reverse-coded items, both the human and model responses have already been
    transformed with 6 - raw_score before entering this function.
    """
    diffs = {domain: [] for domain in DOMAIN_ORDER}
    for idx, item_num in enumerate(TEST_ITEMS):
        domain = ITEM_TO_DOMAIN[item_num]
        h = normalize_raw_score(human_scored_test[idx])
        m = normalize_raw_score(model_scored_test[idx])
        diffs[domain].append(abs(h - m))
    out = {f"aligned_{domain}": round(float(np.mean(vals)), 4) for domain, vals in diffs.items()}
    out["aligned_composite"] = round(float(sum(out[f"aligned_{d}"] for d in DOMAIN_ORDER)), 4)
    return out


def compute_scored_alignment_by_facet(human_scored_test: Sequence[int], model_scored_test: Sequence[int]) -> Dict[str, float]:
    """MPI-style aligned score by each IPIP facet after reverse coding both sides.

    For the validation split, each facet contributes four items from i181-i300.
    """
    diffs = {facet: [] for facet in FACET_ORDER}
    for idx, item_num in enumerate(TEST_ITEMS):
        facet = ITEM_TO_FACET[item_num]
        h = normalize_raw_score(human_scored_test[idx])
        m = normalize_raw_score(model_scored_test[idx])
        diffs[facet].append(abs(h - m))
    return {facet_alignment_col(facet): round(float(np.mean(vals)), 4) for facet, vals in diffs.items()}


def scores_from_result_row(row: Any, side: str = "human") -> List[int]:
    """Return reverse-scored test score array from a result row/dict."""
    getter = row.get if isinstance(row, dict) else row.get

    candidates = [
        f"{side}_scores_scored",
        f"{side}_scores",
    ]
    suffix = str(getter("test_n", "") or "").strip()
    if suffix:
        candidates.extend([
            f"{side}_scores_{suffix}_scored",
            f"{side}_scores_{suffix}",
        ])
    candidates.extend([
        f"{side}_scores_180_scored",
        f"{side}_scores_180",
        f"{side}_scores_120_scored",
        f"{side}_scores_120",
    ])

    raw = None
    for key in candidates:
        raw = getter(key, None)
        if raw is not None and not (isinstance(raw, float) and math.isnan(raw)):
            break

    if isinstance(raw, str):
        try:
            return [normalize_raw_score(x) for x in json.loads(raw)]
        except Exception:
            return []
    if isinstance(raw, list):
        return [normalize_raw_score(x) for x in raw]
    return []

def ensure_facet_alignment_columns(results_df: pd.DataFrame) -> pd.DataFrame:
    """Fill facet-level aligned score columns for old/new partial results."""
    if results_df.empty:
        return results_df
    out = results_df.copy()
    for facet in FACET_ORDER:
        col = facet_alignment_col(facet)
        if col not in out.columns:
            out[col] = np.nan
    for idx, row in out.iterrows():
        missing_any = any(pd.isna(row.get(facet_alignment_col(facet))) for facet in FACET_ORDER)
        if not missing_any:
            continue
        human = scores_from_result_row(row, "human")
        model = scores_from_result_row(row, "model")
        if len(human) != len(TEST_ITEMS) or len(model) != len(TEST_ITEMS):
            continue
        facet_vals = compute_scored_alignment_by_facet(human, model)
        for col, value in facet_vals.items():
            out.at[idx, col] = value
    return out


def check_central_tendency_bias(human_scores: Sequence[int], model_scores: Sequence[int]) -> Dict[str, Any]:
    human = np.array([normalize_raw_score(s) for s in human_scores], dtype=float)
    model = np.array([normalize_raw_score(s) for s in model_scores], dtype=float)
    human_var = float(np.var(human))
    model_var = float(np.var(model))
    ratio = round(model_var / human_var, 4) if human_var > 0 else None
    return {
        "human_mean": round(float(np.mean(human)), 4),
        "model_mean": round(float(np.mean(model)), 4),
        "human_variance": round(human_var, 4),
        "model_variance": round(model_var, 4),
        "variance_ratio": ratio,
        "model_neutral_pct": round(float(np.sum(model == 3) / len(model) * 100), 2),
        "central_bias": "YES" if ratio is not None and ratio < 0.70 else "NO",
    }



def compute_case_metrics(
    row: pd.Series,
    model_scores_test: Sequence[int],
    instructor_key: str,
    target_key: str,
) -> Dict[str, Any]:
    """Compute alignment-focused case metrics.

    The target LLM answers raw Likert values. For validation, both human and
    LLM answers are reverse-coded into personality-direction values first.
    The exported metrics are alignment scores only: OCEAN, facet, and their
    mean composites. General error metrics are intentionally not calculated or exported.
    """
    case_id = str(row.get("case", "unknown"))

    # Raw values are the original questionnaire responses.
    human_test_raw = get_scores(row, TEST_ITEMS)
    model_test_raw = [normalize_raw_score(s) for s in model_scores_test]

    # Scored values are the personality-direction values after applying reverse coding.
    human_test_scored = score_sequence_by_items(human_test_raw, TEST_ITEMS)
    model_test_scored = score_sequence_by_items(model_test_raw, TEST_ITEMS)

    aligned = compute_scored_alignment_by_domain(human_test_scored, model_test_scored)
    facet_aligned = compute_scored_alignment_by_facet(human_test_scored, model_test_scored)

    ocean_vals = [aligned.get(f"aligned_{d}") for d in DOMAIN_ORDER]
    ocean_vals = [float(v) for v in ocean_vals if v is not None and pd.notna(v)]
    facet_vals = [facet_aligned.get(facet_alignment_col(f)) for f in FACET_ORDER]
    facet_vals = [float(v) for v in facet_vals if v is not None and pd.notna(v)]

    # Bias values are retained for auditing only, not for ranking.
    bias = check_central_tendency_bias(human_test_scored, model_test_scored)
    raw_bias = check_central_tendency_bias(human_test_raw, model_test_raw)

    # OCEAN and facet mean scores are computed from raw responses using the official reverse-scoring key.
    # These are not item-alignment scores; they are mean personality scores on the validation split.
    human_ocean = compute_domain_avgs_from_scores(human_test_raw, TEST_KEY, TEST_ITEMS[0])
    model_ocean = compute_domain_avgs_from_scores(model_test_raw, TEST_KEY, TEST_ITEMS[0])
    human_facets = compute_facet_avgs_from_scores(human_test_raw, TEST_KEY, TEST_ITEMS[0])
    model_facets = compute_facet_avgs_from_scores(model_test_raw, TEST_KEY, TEST_ITEMS[0])

    result: Dict[str, Any] = {
        "case": case_id,
        "age": row.get("age", "N/A"),
        "sex": row.get("sex", "N/A"),
        "country": row.get("country", "N/A"),
        "instructor_model": instructor_key,
        "target_model": target_key,
        "model_pair": f"{instructor_key} -> {target_key}",
        "comparison_scale": "reverse_scored_personality_direction",
        **aligned,
        **facet_aligned,
        "aligned_ocean_mean": round(float(np.mean(ocean_vals)), 4) if ocean_vals else None,
        "aligned_facet_mean": round(float(np.mean(facet_vals)), 4) if facet_vals else None,
        **bias,
        "raw_human_mean": raw_bias["human_mean"],
        "raw_model_mean": raw_bias["model_mean"],
        "raw_human_variance": raw_bias["human_variance"],
        "raw_model_variance": raw_bias["model_variance"],
        "raw_variance_ratio": raw_bias["variance_ratio"],
        "raw_model_neutral_pct": raw_bias["model_neutral_pct"],
        "test_n": len(TEST_ITEMS),
        "train_items": TRAIN_RANGE_LABEL,
        "test_items": TEST_RANGE_LABEL,
        "questionnaire_mode": QUESTIONNAIRE_MODE,
        "human_scores": json.dumps(human_test_scored),
        "model_scores": json.dumps(model_test_scored),
        "human_scores_scored": json.dumps(human_test_scored),
        "model_scores_scored": json.dumps(model_test_scored),
        "human_scores_raw": json.dumps(human_test_raw),
        "model_scores_raw": json.dumps(model_test_raw),
        f"human_scores_{TEST_SCORE_SUFFIX}": json.dumps(human_test_scored),
        f"model_scores_{TEST_SCORE_SUFFIX}": json.dumps(model_test_scored),
        f"human_scores_{TEST_SCORE_SUFFIX}_scored": json.dumps(human_test_scored),
        f"model_scores_{TEST_SCORE_SUFFIX}_scored": json.dumps(model_test_scored),
        f"human_scores_{TEST_SCORE_SUFFIX}_raw": json.dumps(human_test_raw),
        f"model_scores_{TEST_SCORE_SUFFIX}_raw": json.dumps(model_test_raw),
    }

    for domain in DOMAIN_ORDER:
        result[f"human_{domain}_avg"] = human_ocean[domain]
        result[f"model_{domain}_avg"] = model_ocean[domain]
        result[f"diff_{domain}_avg"] = round(model_ocean[domain] - human_ocean[domain], 4)
        result[f"absdiff_{domain}_avg"] = round(abs(model_ocean[domain] - human_ocean[domain]), 4)

    facet_absdiffs = []
    for facet in FACET_ORDER:
        short = facet_short_name(facet)
        h_val = human_facets.get(facet)
        m_val = model_facets.get(facet)
        result[f"human_{short}_avg"] = h_val
        result[f"model_{short}_avg"] = m_val
        if h_val is not None and m_val is not None:
            diff_val = round(float(m_val) - float(h_val), 4)
            abs_val = round(abs(diff_val), 4)
            result[f"diff_{short}_avg"] = diff_val
            result[f"absdiff_{short}_avg"] = abs_val
            facet_absdiffs.append(abs_val)
        else:
            result[f"diff_{short}_avg"] = None
            result[f"absdiff_{short}_avg"] = None

    result["ocean_mean_absdiff"] = round(float(np.mean([result[f"absdiff_{d}_avg"] for d in DOMAIN_ORDER])), 4)
    result["facet_mean_absdiff"] = round(float(np.mean(facet_absdiffs)), 4) if facet_absdiffs else None
    return result



def build_statistical_tests(results_df: pd.DataFrame) -> pd.DataFrame:
    """Domain statistical-test export removed by design.

    This validation version is alignment-focused. It no longer creates a
    separate domain statistics file or general error-metric outputs.
    """
    return pd.DataFrame()



def build_model_ranking(results_df: pd.DataFrame) -> pd.DataFrame:
    """Build an alignment-only model ranking table."""
    aggregations = {
        "aligned_composite": "mean",
        "aligned_ocean_mean": "mean",
        "aligned_facet_mean": "mean",
        "ocean_mean_absdiff": "mean",
        "facet_mean_absdiff": "mean",
        "variance_ratio": "mean",
        "model_neutral_pct": "mean",
    }
    ranking = (
        results_df
        .groupby((["prompt_type", "questionnaire_mode"] if "prompt_type" in results_df.columns and "questionnaire_mode" in results_df.columns else ["prompt_type"] if "prompt_type" in results_df.columns else []) + ["instructor_model", "target_model", "model_pair"], as_index=False)
        .agg(aggregations)
        .rename(columns={
            "aligned_composite": "Align",
            "aligned_ocean_mean": "O_mean",
            "aligned_facet_mean": "F_mean",
            "ocean_mean_absdiff": "O_absdiff",
            "facet_mean_absdiff": "F_absdiff",
            "variance_ratio": "VarRatio",
            "model_neutral_pct": "Neutral%",
        })
    )

    ranking = ranking.sort_values(
        by=["Align", "O_mean", "F_mean", "O_absdiff", "F_absdiff"],
        ascending=[True, True, True, True, True],
    ).reset_index(drop=True)
    ranking.insert(0, "Rank", range(1, len(ranking) + 1))
    return ranking


def save_fig(fig: plt.Figure, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)
    return path



def plot_ranking_bar(ranking_df: pd.DataFrame, plots_dir: Path) -> Path:
    fig, ax = plt.subplots(figsize=(max(10, len(ranking_df) * 0.9), 5))
    labels = ranking_df["model_pair"].tolist()
    values = ranking_df["Align"].astype(float).tolist()
    x = np.arange(len(labels))
    ax.bar(x, values)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=8)
    ax.set_ylabel("Mean alignment score (lower = better)")
    ax.set_title("Model-pair ranking by alignment score")
    ax.grid(axis="y", alpha=0.3)
    for i, val in enumerate(values):
        ax.text(i, val + 0.02, f"{val:.3f}", ha="center", va="bottom", fontsize=8)
    return save_fig(fig, plots_dir / "ranking_alignment.png")



def plot_facet_mean_bar(ranking_df: pd.DataFrame, plots_dir: Path) -> Path:
    fig, ax = plt.subplots(figsize=(max(10, len(ranking_df) * 0.9), 5))
    labels = ranking_df["model_pair"].tolist()
    values = ranking_df["F_mean"].astype(float).tolist()
    x = np.arange(len(labels))
    ax.bar(x, values)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=8)
    ax.set_ylabel("Mean facet alignment score (lower = better)")
    ax.set_title("Model-pair ranking by facet mean alignment")
    ax.grid(axis="y", alpha=0.3)
    for i, val in enumerate(values):
        ax.text(i, val + 0.02, f"{val:.3f}", ha="center", va="bottom", fontsize=8)
    return save_fig(fig, plots_dir / "ranking_facet_mean_alignment.png")


def plot_ocean_absdiff_heatmap(ranking_df: pd.DataFrame, results_df: pd.DataFrame, plots_dir: Path) -> Path:
    ordered_pairs = ranking_df["model_pair"].tolist()
    matrix = []
    for pair in ordered_pairs:
        pair_df = results_df[results_df["model_pair"] == pair]
        matrix.append([float(pair_df[f"absdiff_{d}_avg"].mean()) for d in DOMAIN_ORDER])
    arr = np.array(matrix)

    fig, ax = plt.subplots(figsize=(8, max(4, len(ordered_pairs) * 0.45)))
    im = ax.imshow(arr, aspect="auto")
    ax.set_xticks(np.arange(len(DOMAIN_ORDER)))
    ax.set_xticklabels([DOMAIN_FULL_NAMES[d] for d in DOMAIN_ORDER], rotation=25, ha="right")
    ax.set_yticks(np.arange(len(ordered_pairs)))
    ax.set_yticklabels(ordered_pairs, fontsize=8)
    ax.set_title("Mean absolute OCEAN average difference by model pair")
    cbar = fig.colorbar(im, ax=ax)
    cbar.set_label("|model avg - human avg|")
    for i in range(arr.shape[0]):
        for j in range(arr.shape[1]):
            ax.text(j, i, f"{arr[i, j]:.2f}", ha="center", va="center", fontsize=8)
    return save_fig(fig, plots_dir / "ocean_absdiff_heatmap.png")


def plot_pair_ocean_bars(pair_df: pd.DataFrame, pair_slug: str, plots_dir: Path) -> Path:
    human = [float(pair_df[f"human_{d}_avg"].mean()) for d in DOMAIN_ORDER]
    model = [float(pair_df[f"model_{d}_avg"].mean()) for d in DOMAIN_ORDER]
    labels = [DOMAIN_FULL_NAMES[d] for d in DOMAIN_ORDER]
    x = np.arange(len(DOMAIN_ORDER))
    width = 0.36

    fig, ax = plt.subplots(figsize=(9, 5))
    ax.bar(x - width / 2, human, width, label="Human")
    ax.bar(x + width / 2, model, width, label="Model")
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=20, ha="right")
    ax.set_ylabel("Average score after reverse scoring")
    ax.set_ylim(1, 5)
    ax.set_title(f"OCEAN averages — {pair_df['model_pair'].iloc[0]}")
    ax.legend()
    ax.grid(axis="y", alpha=0.3)
    for i, value in enumerate(human):
        ax.text(i - width / 2, value + 0.03, f"{value:.2f}", ha="center", fontsize=8)
    for i, value in enumerate(model):
        ax.text(i + width / 2, value + 0.03, f"{value:.2f}", ha="center", fontsize=8)
    return save_fig(fig, plots_dir / f"{pair_slug}_ocean_human_vs_model.png")


def plot_pair_score_distribution(pair_df: pd.DataFrame, pair_slug: str, plots_dir: Path) -> Path:
    human_all: List[int] = []
    model_all: List[int] = []
    for _, row in pair_df.iterrows():
        human_all.extend(scores_from_result_row(row, "human"))
        model_all.extend(scores_from_result_row(row, "model"))

    human_counts = [human_all.count(i) for i in range(1, 6)]
    model_counts = [model_all.count(i) for i in range(1, 6)]
    x = np.arange(1, 6)
    width = 0.36

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.bar(x - width / 2, human_counts, width, label="Human")
    ax.bar(x + width / 2, model_counts, width, label="Model")
    ax.set_xticks(x)
    ax.set_xlabel("Reverse-scored personality-direction score")
    ax.set_ylabel(f"Count across {TEST_RANGE_LABEL}")
    ax.set_title(f"Reverse-scored score distribution — {pair_df['model_pair'].iloc[0]}")
    ax.legend()
    ax.grid(axis="y", alpha=0.3)
    return save_fig(fig, plots_dir / f"{pair_slug}_scored_score_distribution.png")


def plot_pair_violin(pair_df: pd.DataFrame, pair_slug: str, plots_dir: Path) -> Path:
    diffs = {domain: [] for domain in DOMAIN_ORDER}
    for _, row in pair_df.iterrows():
        human = scores_from_result_row(row, "human")
        model = scores_from_result_row(row, "model")
        for idx, item_num in enumerate(TEST_ITEMS):
            domain = ITEM_TO_DOMAIN[item_num]
            diffs[domain].append(normalize_raw_score(human[idx]) - normalize_raw_score(model[idx]))

    data = [diffs[d] for d in DOMAIN_ORDER]
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.violinplot(data, positions=np.arange(1, len(DOMAIN_ORDER) + 1), showmeans=True, showmedians=True)
    ax.axhline(0, linestyle="--", linewidth=1)
    ax.set_xticks(np.arange(1, len(DOMAIN_ORDER) + 1))
    ax.set_xticklabels([DOMAIN_FULL_NAMES[d] for d in DOMAIN_ORDER], rotation=20, ha="right")
    ax.set_ylabel("Human reverse-scored value minus model reverse-scored value")
    ax.set_title(f"Distribution of score differences — {pair_df['model_pair'].iloc[0]}")
    ax.grid(axis="y", alpha=0.3)
    return save_fig(fig, plots_dir / f"{pair_slug}_difference_violin.png")


def plot_case_line(case_id: str, human: Sequence[int], model: Sequence[int], pair_slug: str, plots_dir: Path) -> Path:
    fig, ax = plt.subplots(figsize=(14, 4))
    x = np.array(TEST_ITEMS)
    ax.plot(x, human, linewidth=0.9, label="Human")
    ax.plot(x, model, linewidth=0.9, label="Model")
    ax.set_ylim(0.5, 5.5)
    ax.set_yticks([1, 2, 3, 4, 5])
    ax.set_xlabel(f"Item number ({TEST_RANGE_LABEL})")
    ax.set_ylabel("Reverse-scored value")
    ax.set_title(f"Case {case_id} — Human vs model responses after reverse scoring")
    ax.grid(axis="y", alpha=0.3)
    ax.legend()
    return save_fig(fig, plots_dir / f"{pair_slug}_case_{case_id}_line.png")


# =============================================================================
# Export helpers
# =============================================================================

def facet_alignment_columns() -> List[str]:
    return [facet_alignment_col(facet) for facet in FACET_ORDER]



def build_pair_detail_sheet(pair_df: pd.DataFrame) -> pd.DataFrame:
    """Backward-compatible helper: detailed case-level alignment sheet for one pair."""
    return build_alignment_scores_sheet(pair_df)



def build_alignment_scores_sheet(results_df: pd.DataFrame) -> pd.DataFrame:
    """Build the main Alignment Scores sheet.

    Compact alignment-only columns:
    Pair, Case, O/C/E/A/N, O_mean, all facet alignment scores, F_mean.
    The AVG row appears at the bottom of each model-pair block.
    """
    if results_df.empty:
        return pd.DataFrame()

    results_df = ensure_facet_alignment_columns(results_df)
    rows: List[Dict[str, Any]] = []

    if "prompt_type" in results_df.columns and "questionnaire_mode" in results_df.columns:
        group_cols = ["prompt_type", "questionnaire_mode", "instructor_model", "target_model"]
    elif "prompt_type" in results_df.columns:
        group_cols = ["prompt_type", "instructor_model", "target_model"]
    else:
        group_cols = ["instructor_model", "target_model"]

    for group_key, pair_df in results_df.groupby(group_cols, dropna=False):
        if len(group_cols) == 4:
            prompt_type, questionnaire_mode, instructor, target = group_key
        elif len(group_cols) == 3:
            prompt_type, instructor, target = group_key
            questionnaire_mode = "120-180"
        else:
            prompt_type, questionnaire_mode, instructor, target = "1", "120-180", group_key[0], group_key[1]
        pair_label = f"{instructor} -> {target} [P{prompt_type} | Q{questionnaire_mode}]"
        pair_df = pair_df.copy().sort_values(by=["case"])

        for _, r in pair_df.iterrows():
            row: Dict[str, Any] = {
                "Pair": pair_label,
                "Case": r.get("case"),
            }
            ocean_vals = []
            for domain in DOMAIN_ORDER:
                val = r.get(f"aligned_{domain}")
                row[domain] = val
                if pd.notna(val):
                    ocean_vals.append(float(val))
            row["O_mean"] = round(float(np.mean(ocean_vals)), 4) if ocean_vals else None

            facet_vals = []
            for facet in FACET_ORDER:
                col = facet_alignment_col(facet)
                val = r.get(col)
                short = facet_short_name(facet)
                row[short] = val
                if pd.notna(val):
                    facet_vals.append(float(val))
            row["F_mean"] = round(float(np.mean(facet_vals)), 4) if facet_vals else None
            rows.append(row)

        avg: Dict[str, Any] = {"Pair": pair_label, "Case": "AVG"}
        numeric_cols = [*DOMAIN_ORDER, "O_mean", *[facet_short_name(f) for f in FACET_ORDER], "F_mean"]
        for col in numeric_cols:
            vals = pd.to_numeric(pd.Series([row.get(col) for row in rows if row.get("Pair") == pair_label and row.get("Case") != "AVG"]), errors="coerce")
            avg[col] = round(float(vals.mean()), 4) if vals.notna().any() else None
        rows.append(avg)

        # blank separator row between model pairs for readability
        rows.append({col: "" for col in ["Pair", "Case", *DOMAIN_ORDER, "O_mean", *[facet_short_name(f) for f in FACET_ORDER], "F_mean"]})

    if rows and all(v == "" for v in rows[-1].values()):
        rows = rows[:-1]
    return pd.DataFrame(rows)



def build_combination_comparison(results_df: pd.DataFrame) -> pd.DataFrame:
    """Cumulative/current model-combo comparison with alignment-only columns."""
    if results_df.empty:
        return pd.DataFrame()
    results_df = ensure_facet_alignment_columns(results_df)
    rows: List[Dict[str, Any]] = []

    if "prompt_type" in results_df.columns and "questionnaire_mode" in results_df.columns:
        group_cols = ["prompt_type", "questionnaire_mode", "instructor_model", "target_model"]
    elif "prompt_type" in results_df.columns:
        group_cols = ["prompt_type", "instructor_model", "target_model"]
    else:
        group_cols = ["instructor_model", "target_model"]

    for group_key, pair_df in results_df.groupby(group_cols, dropna=False):
        if len(group_cols) == 4:
            prompt_type, questionnaire_mode, instructor, target = group_key
        elif len(group_cols) == 3:
            prompt_type, instructor, target = group_key
            questionnaire_mode = "120-180"
        else:
            prompt_type, questionnaire_mode, instructor, target = "1", "120-180", group_key[0], group_key[1]
        row: Dict[str, Any] = {
            "P": prompt_type,
            "Q": questionnaire_mode,
            "Instr": instructor,
            "Target": target,
            "Pair": f"{instructor} -> {target} [P{prompt_type} | Q{questionnaire_mode}]",
            "N": int(pair_df["case"].nunique()) if "case" in pair_df.columns else len(pair_df),
        }

        ocean_vals = []
        for domain in DOMAIN_ORDER:
            val = round(float(pd.to_numeric(pair_df.get(f"aligned_{domain}"), errors="coerce").mean()), 4)
            row[domain] = val
            if pd.notna(val):
                ocean_vals.append(val)
        row["O_mean"] = round(float(np.mean(ocean_vals)), 4) if ocean_vals else None

        facet_vals = []
        for facet in FACET_ORDER:
            col = facet_alignment_col(facet)
            val = round(float(pd.to_numeric(pair_df.get(col), errors="coerce").mean()), 4)
            short = facet_short_name(facet)
            row[short] = val
            if pd.notna(val):
                facet_vals.append(val)
        row["F_mean"] = round(float(np.mean(facet_vals)), 4) if facet_vals else None

        rows.append(row)

    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values(by=["O_mean", "F_mean"], ascending=[True, True]).reset_index(drop=True)
        out.insert(0, "Rank", range(1, len(out) + 1))
    return out


def load_cumulative_case_results(current_output_dir: Optional[Path] = None) -> pd.DataFrame:
    """Load successful partial results across validation_outputs/run_* and de-duplicate by case/model pair.

    This is what keeps the cross-model comparison sheet updating across separate runs.
    Latest result wins if the same case/instructor/target appears more than once.
    """
    records: List[Tuple[float, int, Dict[str, Any]]] = []
    order = 0
    for partial_path in sorted(OUTPUT_ROOT.glob(f"*/{PARTIAL_RESULTS_FILENAME}")):
        try:
            mtime = partial_path.stat().st_mtime
            run_id = partial_path.parent.name
            for line in partial_path.read_text(encoding="utf-8").splitlines():
                if not line.strip():
                    continue
                try:
                    payload = json.loads(line)
                except Exception:
                    continue
                payload["source_run"] = run_id
                payload["source_output_dir"] = str(partial_path.parent)
                records.append((mtime, order, payload))
                order += 1
        except Exception:
            continue

    deduped: Dict[Tuple[str, str, str, str, str], Tuple[float, int, Dict[str, Any]]] = {}
    for mtime, order_idx, payload in records:
        key = result_unique_key(payload)
        current = deduped.get(key)
        if current is None or (mtime, order_idx) >= (current[0], current[1]):
            deduped[key] = (mtime, order_idx, payload)

    rows = [payload for _, _, payload in sorted(deduped.values(), key=lambda item: (item[0], item[1]))]
    if not rows:
        return pd.DataFrame()
    return ensure_facet_alignment_columns(pd.DataFrame(rows))


def build_mean_deviation_sheet(results_df: pd.DataFrame) -> pd.DataFrame:
    """Long-format human-vs-LLM mean score deviations for OCEAN traits and facets.

    This is separate from alignment scores. Alignment scores compare item-by-item;
    mean deviation compares the resulting mean trait/facet scores.
    """
    if results_df.empty:
        return pd.DataFrame()

    rows: List[Dict[str, Any]] = []
    if "prompt_type" in results_df.columns and "questionnaire_mode" in results_df.columns:
        group_cols = ["prompt_type", "questionnaire_mode", "instructor_model", "target_model"]
    elif "prompt_type" in results_df.columns:
        group_cols = ["prompt_type", "instructor_model", "target_model"]
    else:
        group_cols = ["instructor_model", "target_model"]

    for group_key, pair_df in results_df.groupby(group_cols, dropna=False):
        if len(group_cols) == 4:
            prompt_type, questionnaire_mode, instructor, target = group_key
        elif len(group_cols) == 3:
            prompt_type, instructor, target = group_key
            questionnaire_mode = "120-180"
        else:
            prompt_type, questionnaire_mode, instructor, target = "1", "120-180", group_key[0], group_key[1]
        pair_label = f"{instructor} -> {target} [P{prompt_type} | Q{questionnaire_mode}]"
        pair_df = pair_df.copy().sort_values(by=["case"])

        case_rows: List[Dict[str, Any]] = []
        for _, r in pair_df.iterrows():
            case_id = r.get("case")
            for domain in DOMAIN_ORDER:
                case_rows.append({
                    "Pair": pair_label,
                    "Case": case_id,
                    "Level": "OCEAN",
                    "Name": domain,
                    "H": r.get(f"human_{domain}_avg"),
                    "LLM": r.get(f"model_{domain}_avg"),
                    "Diff": r.get(f"diff_{domain}_avg"),
                    "Abs": r.get(f"absdiff_{domain}_avg"),
                })

            for facet in FACET_ORDER:
                short = facet_short_name(facet)
                case_rows.append({
                    "Pair": pair_label,
                    "Case": case_id,
                    "Level": "Facet",
                    "Name": short,
                    "H": r.get(f"human_{short}_avg"),
                    "LLM": r.get(f"model_{short}_avg"),
                    "Diff": r.get(f"diff_{short}_avg"),
                    "Abs": r.get(f"absdiff_{short}_avg"),
                })

        rows.extend(case_rows)

        for level, names in [("OCEAN", DOMAIN_ORDER), ("Facet", [facet_short_name(f) for f in FACET_ORDER])]:
            for name in names:
                subset = [x for x in case_rows if x["Level"] == level and x["Name"] == name]
                avg_row: Dict[str, Any] = {"Pair": pair_label, "Case": "AVG", "Level": level, "Name": name}
                for col in ["H", "LLM", "Diff", "Abs"]:
                    vals = pd.to_numeric(pd.Series([x.get(col) for x in subset]), errors="coerce")
                    avg_row[col] = round(float(vals.mean()), 4) if vals.notna().any() else None
                rows.append(avg_row)

        rows.append({"Pair": "", "Case": "", "Level": "", "Name": "", "H": "", "LLM": "", "Diff": "", "Abs": ""})

    if rows and all(v == "" for v in rows[-1].values()):
        rows = rows[:-1]
    return pd.DataFrame(rows)


def write_excel_with_plots(
    results_df: pd.DataFrame,
    ranking_df: pd.DataFrame,
    cumulative_comparison_df: pd.DataFrame,
    output_dir: Path,
    plots: Dict[str, Path],
    logger: logging.Logger,
) -> Path:
    """Write a compact workbook:
    - Alignment Scores: case-level OCEAN/facet item-alignment with AVG rows
    - Mean Dev: human-vs-LLM mean trait/facet score deviation with AVG rows
    - Combo Compare: cumulative model-pair comparison that updates across runs
    - Plots: all plots embedded together
    """
    excel_path = output_dir / "alignment_scores.xlsx"

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        alignment_df = build_alignment_scores_sheet(results_df)
        alignment_df.to_excel(writer, sheet_name="Alignment Scores", index=False)

        mean_dev_df = build_mean_deviation_sheet(results_df)
        mean_dev_df.to_excel(writer, sheet_name="Mean Dev", index=False)

        # This is cumulative across all trial folders, so it keeps updating as new model combinations are run.
        cumulative_comparison_df.to_excel(writer, sheet_name="Combo Compare", index=False)

        # Keep current-run comparison as CSV, but not as another Excel sheet to avoid clutter.
        # Keep prompts in JSONL, not Excel, because they are too wide for quick reading.

    if load_workbook is not None and XLImage is not None:
        try:
            wb = load_workbook(excel_path)

            # Basic readability: freeze top rows and make common numeric columns reasonably narrow.
            for sheet_name in ["Alignment Scores", "Mean Dev", "Combo Compare"]:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    ws.freeze_panes = "A2"
                    for col_cells in ws.columns:
                        header = str(col_cells[0].value or "")
                        width = 8
                        if header in {"Pair", "Instr", "Target"}:
                            width = 22
                        elif header in {"Case", "Level", "Name"}:
                            width = 10
                        elif header in {"Rank", "N"}:
                            width = 6
                        ws.column_dimensions[col_cells[0].column_letter].width = width

            if "Plots" in wb.sheetnames:
                del wb["Plots"]
            sheet = wb.create_sheet("Plots")
            row = 1
            for label, path in plots.items():
                if not Path(path).exists():
                    continue
                sheet.cell(row=row, column=1, value=label)
                img = XLImage(str(path))
                img.width = 900
                img.height = 430
                sheet.add_image(img, f"A{row + 1}")
                row += 25
            wb.save(excel_path)
        except Exception as e:
            logger.warning(f"Could not format workbook or embed plots in Excel: {e}")

    return excel_path



def write_jsonl_prompts(results: List[Dict[str, Any]], output_dir: Path) -> Path:
    path = output_dir / "prompts_scores.jsonl"
    with path.open("w", encoding="utf-8") as f:
        for row in results:
            payload = {
                "case": row.get("case"),
                "instructor_model": row.get("instructor_model"),
                "target_model": row.get("target_model"),
                "prompt_type": row.get("prompt_type"),
                "instructor_prompt": row.get("instructor_prompt"),
                "validation_system_prompt": row.get("validation_system_prompt"),
                "questionnaire_mode": row.get("questionnaire_mode"),
                "train_items": row.get("train_items"),
                "test_items": row.get("test_items"),
                "human_scores_scored": row.get("human_scores_scored"),
                "model_scores_scored": row.get("model_scores_scored"),
                "human_scores_raw": row.get("human_scores_raw"),
                "model_scores_raw": row.get("model_scores_raw"),
            }
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")
    return path



# =============================================================================
# Resumable run helpers
# =============================================================================

def load_json_file(path: Path) -> Optional[Dict[str, Any]]:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def save_json_file(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)


def choose_output_dir_for_run(
    args_output_dir: Optional[str],
    timestamp: str,
    selection: Optional[Dict[str, List[str]]] = None,
) -> Tuple[Path, bool]:
    """Choose a new model-combo/trial output directory or resume the latest incomplete run."""
    if args_output_dir:
        return Path(args_output_dir).expanduser().resolve(), False

    state = load_json_file(RESUME_STATE_FILE)
    if state and not state.get("completed", False):
        previous_dir = Path(state.get("output_dir", ""))
        if previous_dir.exists():
            print("\n" + "=" * 86)
            print("PREVIOUS VALIDATION RUN FOUND")
            print("=" * 86)
            print(f"Previous output folder: {previous_dir}")
            print(f"Started at             : {state.get('started_at', 'unknown')}")
            print(f"Instructor model(s)    : {', '.join(state.get('selection', {}).get('instructors', []))}")
            print(f"Target model(s)        : {', '.join(state.get('selection', {}).get('targets', []))}")
            print(f"Prompt type            : {state.get('selection', {}).get('prompt_type', '1')}")
            print(f"Questionnaire mode     : {state.get('selection', {}).get('questionnaire_mode', '120-180')}")
            print("\nPress 1 to continue from where it stopped")
            print("Press 2 to start a new run")
            choice = input("Choice: ").strip()
            if choice == "1":
                print("Continuing previous run. Previous successful case results will be merged into the final outputs.")
                return previous_dir, True

    if selection is None:
        return OUTPUT_ROOT / f"trial_{next_trial_number(OUTPUT_ROOT):03d}_model_combo_{timestamp}", False
    return make_trial_output_dir(selection, timestamp), False



def save_run_manifest(
    output_dir: Path,
    selection: Dict[str, List[str]],
    csv_path: Path,
    max_participants: Optional[int],
    batch_size: int,
    case_plots: bool,
    completed: bool = False,
) -> None:
    manifest = {
        "prompt_version": PROMPT_VERSION,
        "selection": selection,
        "csv_path": str(csv_path),
        "max_participants": max_participants,
        "batch_size": batch_size,
        "case_plots": case_plots,
        "output_dir": str(output_dir),
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "completed": completed,
    }
    save_json_file(output_dir / RUN_MANIFEST_FILENAME, manifest)
    save_json_file(RESUME_STATE_FILE, manifest)


def mark_resume_state_completed(output_dir: Path, completed: bool) -> None:
    manifest_path = output_dir / RUN_MANIFEST_FILENAME
    manifest = load_json_file(manifest_path) or {}
    manifest["completed"] = completed
    manifest["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
    save_json_file(manifest_path, manifest)
    save_json_file(RESUME_STATE_FILE, manifest)


def result_unique_key(row: Dict[str, Any]) -> Tuple[str, str, str, str, str]:
    return (
        str(row.get("case")),
        str(row.get("instructor_model")),
        str(row.get("target_model")),
        str(row.get("prompt_type", "1")),
        str(row.get("questionnaire_mode", "120-180")),
    )

def append_partial_result(output_dir: Path, result: Dict[str, Any]) -> None:
    partial_path = output_dir / PARTIAL_RESULTS_FILENAME
    with partial_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(result, ensure_ascii=False) + "\n")


def load_partial_results(output_dir: Path) -> List[Dict[str, Any]]:
    """Load previous successful case results for a resumable run, de-duplicated by case/model pair."""
    partial_path = output_dir / PARTIAL_RESULTS_FILENAME
    rows: List[Dict[str, Any]] = []
    if partial_path.exists():
        for line in partial_path.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            try:
                rows.append(json.loads(line))
            except Exception:
                continue

    # Fallback: if a completed CSV exists but partial_results.jsonl does not, preserve what can be recovered.
    csv_path = output_dir / "all_case_results.csv"
    if not rows and csv_path.exists():
        try:
            recovered = pd.read_csv(csv_path).to_dict(orient="records")
            rows.extend(recovered)
        except Exception:
            pass

    deduped: Dict[Tuple[str, str, str, str, str], Dict[str, Any]] = {}
    for row in rows:
        deduped[result_unique_key(row)] = row
    return list(deduped.values())


def expected_result_keys(df: pd.DataFrame, selection: Dict[str, List[str]]) -> set[Tuple[str, str, str, str, str]]:
    keys = set()
    prompt_type = str(selection.get("prompt_type", "1"))
    questionnaire_mode = str(selection.get("questionnaire_mode", "120-180"))
    for instructor_key in selection["instructors"]:
        for target_key in selection["targets"]:
            for _, row in df.iterrows():
                keys.add((str(row.get("case", "unknown")), instructor_key, target_key, prompt_type, questionnaire_mode))
    return keys

# =============================================================================
# Main validation loop
# =============================================================================

def find_default_csv() -> Path:
    candidates = [
        SCRIPT_DIR / "IPIP_NEO_300.csv",
        SCRIPT_DIR / "prosocial_antisocial_ipip300_answers.csv",
    ]
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError(
        "Could not find IPIP_NEO_300.csv or prosocial_antisocial_ipip300_answers.csv beside the script. "
        "Use --csv to provide a path."
    )


def load_data(csv_path: Path, max_participants: Optional[int]) -> pd.DataFrame:
    df = pd.read_csv(csv_path)
    if "case" not in df.columns:
        if "profile_name" in df.columns:
            df = df.rename(columns={"profile_name": "case"})
        else:
            df["case"] = [str(i + 1) for i in range(len(df))]
    for optional in ["age", "sex", "country"]:
        if optional not in df.columns:
            df[optional] = "N/A"
    df["case"] = df["case"].astype(str)

    missing_cols = [col for col in ALL_COLS if col not in df.columns]
    if missing_cols:
        raise ValueError(f"CSV is missing required item columns: {missing_cols[:10]} ... total={len(missing_cols)}")

    if max_participants is not None:
        df = df.head(max_participants)
    return df


def process_case_pair(
    row: pd.Series,
    instructor_key: str,
    target_key: str,
    logger: logging.Logger,
    batch_size: int,
    prompt_type: str = "1",
) -> Dict[str, Any]:
    case_id = str(row.get("case", "unknown"))

    instructor_prompt = generate_instructor_prompt(row, instructor_key, logger, prompt_type=prompt_type)
    calibration = build_training_calibration(row)
    few_shot = build_few_shot_train(row, prompt_type=prompt_type)
    system_prompt = build_validation_system_prompt(
        instructor_prompt=instructor_prompt,
        calibration_prompt=calibration,
        few_shot_prompt=few_shot,
        instructor_key=instructor_key,
        target_key=target_key,
        case_id=case_id,
        logger=logger,
        prompt_type=prompt_type,
    )

    scores_test = run_target_batches(
        row=row,
        system_prompt=system_prompt,
        target_key=target_key,
        instructor_key=instructor_key,
        logger=logger,
        batch_size=batch_size,
        prompt_type=prompt_type,
    )

    result = compute_case_metrics(
        row=row,
        model_scores_test=scores_test,
        instructor_key=instructor_key,
        target_key=target_key,
    )
    result["prompt_type"] = str(prompt_type)
    result["questionnaire_mode"] = QUESTIONNAIRE_MODE
    result["model_pair"] = f"{instructor_key} -> {target_key} [P{prompt_type} | Q{QUESTIONNAIRE_MODE}]"
    result["instructor_prompt"] = instructor_prompt
    result["validation_system_prompt"] = system_prompt
    result["training_calibration_prompt"] = calibration
    result["few_shot_prompt"] = few_shot
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Run multi-model IPIP-NEO-300 validation.")
    parser.add_argument("--csv", type=str, default=None, help="Path to IPIP_NEO_300.csv or equivalent.")
    parser.add_argument("--max-participants", type=str, default=None, help="Number of participants, or 'all'. If omitted, asks interactively.")
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE, help="Target item batch size. Default: 30.")
    parser.add_argument("--output-dir", type=str, default=None, help="Output directory. Default: validation_outputs/run_TIMESTAMP")
    parser.add_argument("--case-plots", action="store_true", help="Generate line plots for every case and model pair.")
    args = parser.parse_args()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # First handle resume state. If this is a new run, choose models before creating the
    # output folder so the folder name can contain the model-combo and trial number.
    resume_probe_dir, resume_previous = choose_output_dir_for_run(args.output_dir, timestamp, selection=None)
    manifest = load_json_file(resume_probe_dir / RUN_MANIFEST_FILENAME) if resume_previous else None

    if resume_previous and manifest and manifest.get("selection"):
        output_dir = resume_probe_dir
        selection = manifest["selection"]
        print("\nResuming with previous model choice:")
        print(f"Instructor model(s): {', '.join(selection.get('instructors', []))}")
        print(f"Target model(s)    : {', '.join(selection.get('targets', []))}")
        print(f"Prompt type        : {selection.get('prompt_type', '1')} - {prompt_type_description(str(selection.get('prompt_type', '1')))}")
        print(f"Questionnaire mode : {selection.get('questionnaire_mode', '120-180')} - {questionnaire_mode_description(str(selection.get('questionnaire_mode', '120-180')))}")
    else:
        selection = choose_models_interactively()
        output_dir = Path(args.output_dir).expanduser().resolve() if args.output_dir else make_trial_output_dir(selection, timestamp)

    output_dir.mkdir(parents=True, exist_ok=True)
    plots_dir = output_dir / "plots"
    checkpoints_dir = output_dir / "checkpoints"
    checkpoints_dir.mkdir(exist_ok=True)

    logger = setup_logging(output_dir)
    manifest = load_json_file(output_dir / RUN_MANIFEST_FILENAME) if resume_previous else None

    # Configure the selected questionnaire split before data loading and processing.
    # Prompt Type 2 now also respects this split; for 180-120 it uses the same FPS wording adapted to the selected item range.
    configure_questionnaire_split(str(selection.get("questionnaire_mode", "120-180")))

    validate_environment(selection)

    if resume_previous and manifest:
        max_participants = manifest.get("max_participants", DEFAULT_MAX_PARTICIPANTS)
        batch_size = int(manifest.get("batch_size", args.batch_size))
        case_plots = bool(manifest.get("case_plots", args.case_plots))
        csv_path = Path(manifest.get("csv_path")).expanduser().resolve() if manifest.get("csv_path") else (Path(args.csv).expanduser().resolve() if args.csv else find_default_csv())
        logger.info("Resume mode: using previous CSV, participant count, batch size, and model selection.")
    else:
        if args.max_participants is None:
            max_participants = ask_max_participants(DEFAULT_MAX_PARTICIPANTS)
        else:
            max_participants = None if args.max_participants.lower() == "all" else int(args.max_participants)
        batch_size = args.batch_size
        case_plots = args.case_plots
        csv_path = Path(args.csv).expanduser().resolve() if args.csv else find_default_csv()

    df = load_data(csv_path, max_participants=max_participants)

    save_run_manifest(
        output_dir=output_dir,
        selection=selection,
        csv_path=csv_path,
        max_participants=max_participants,
        batch_size=batch_size,
        case_plots=case_plots,
        completed=False,
    )

    logger.info(f"Loaded {len(df)} participants from {csv_path}")
    logger.info(f"Instructor models: {selection['instructors']}")
    logger.info(f"Target models: {selection['targets']}")
    logger.info(f"Prompt type: {selection.get('prompt_type', '1')} - {prompt_type_description(str(selection.get('prompt_type', '1')))}")
    logger.info(f"Questionnaire mode: {selection.get('questionnaire_mode', '120-180')} - {questionnaire_mode_description(str(selection.get('questionnaire_mode', '120-180')))}")
    logger.info(f"Training items: {TRAIN_RANGE_LABEL}; validation items: {TEST_RANGE_LABEL}")
    logger.info(f"Batch size: {batch_size}")
    logger.info("Validation comparison scale: reverse-scored personality-direction values for both human and LLM scores")

    all_results: List[Dict[str, Any]] = load_partial_results(output_dir)
    existing_keys = {result_unique_key(r) for r in all_results}
    if all_results:
        logger.info(f"Loaded {len(all_results)} previous successful case results from partial storage.")

    for instructor_key in selection["instructors"]:
        for target_key in selection["targets"]:
            pair_slug = slugify(f"{instructor_key}__to__{target_key}__p{selection.get('prompt_type', '1')}__q{selection.get('questionnaire_mode', '120-180').replace('-', 'to')}")
            checkpoint_file = checkpoints_dir / f"done_{pair_slug}.txt"
            done = set()
            if checkpoint_file.exists():
                done = set(line.strip() for line in checkpoint_file.read_text(encoding="utf-8").splitlines() if line.strip())
                logger.info(f"Resuming pair {instructor_key} -> {target_key} | already checkpointed: {len(done)}")

            # The partial results file is the source of truth for final merged outputs.
            # If a case exists in partial storage, treat it as done even if checkpoint text is missing.
            done.update({
                case for case, instr, targ, ptype, qmode in existing_keys
                if instr == instructor_key and targ == target_key and ptype == str(selection.get("prompt_type", "1")) and qmode == str(selection.get("questionnaire_mode", "120-180"))
            })

            logger.info("=" * 86)
            logger.info(f"Starting model pair: {instructor_key} -> {target_key}")
            logger.info("=" * 86)

            for i, (_, row) in enumerate(df.iterrows(), start=1):
                case_id = str(row.get("case", "unknown"))
                if case_id in done:
                    logger.info(f"Skipping case {case_id} for {pair_slug} (already completed)")
                    continue

                logger.info(f"[{i}/{len(df)}] case {case_id} | {instructor_key} -> {target_key}")
                try:
                    result = process_case_pair(
                        row=row,
                        instructor_key=instructor_key,
                        target_key=target_key,
                        logger=logger,
                        batch_size=batch_size,
                        prompt_type=str(selection.get("prompt_type", "1")),
                    )
                    key = result_unique_key(result)
                    if key not in existing_keys:
                        all_results.append(result)
                        existing_keys.add(key)
                        append_partial_result(output_dir, result)

                    with checkpoint_file.open("a", encoding="utf-8") as f:
                        f.write(case_id + "\n")

                    logger.info(
                        f"  OK case {case_id} | pair={instructor_key} -> {target_key} | "
                        f"aligned={result['aligned_composite']} | "
                        f"O_mean={result.get('aligned_ocean_mean')} | F_mean={result.get('aligned_facet_mean')} | "
                        f"scale=reverse-scored"
                    )

                    if case_plots:
                        plot_case_line(
                            case_id=case_id,
                            human=json.loads(result["human_scores_scored"]),
                            model=json.loads(result["model_scores_scored"]),
                            pair_slug=pair_slug,
                            plots_dir=plots_dir,
                        )

                except KeyboardInterrupt:
                    logger.warning("Interrupted by user. Successful cases so far have been saved to partial_results.jsonl.")
                    mark_resume_state_completed(output_dir, completed=False)
                    raise
                except Exception as e:
                    logger.error(f"  FAILED case {case_id} | pair={pair_slug} | {repr(e)}")
                    time.sleep(1.0)
                    continue

    if not all_results:
        logger.warning("No results were available. Nothing to export yet.")
        mark_resume_state_completed(output_dir, completed=False)
        return

    results_df = ensure_facet_alignment_columns(pd.DataFrame(all_results))
    ranking_df = build_model_ranking(results_df)
    current_comparison_df = build_combination_comparison(results_df)

    cumulative_results_df = load_cumulative_case_results(current_output_dir=output_dir)
    if cumulative_results_df.empty:
        cumulative_results_df = results_df.copy()
    cumulative_comparison_df = build_combination_comparison(cumulative_results_df)

    # Save only alignment-focused summary files. Case-matrix and domain-stats exports are intentionally omitted.
    ranking_csv = output_dir / "model_combo_ranking.csv"
    current_comparison_csv = output_dir / "combo_summary_current.csv"
    alignment_csv = output_dir / "alignment_scores_current.csv"
    mean_dev_csv = output_dir / "mean_deviation_current.csv"
    build_alignment_scores_sheet(results_df).to_csv(alignment_csv, index=False)
    build_mean_deviation_sheet(results_df).to_csv(mean_dev_csv, index=False)
    ranking_df.to_csv(ranking_csv, index=False)
    current_comparison_df.to_csv(current_comparison_csv, index=False)
    cumulative_comparison_df.to_csv(CUMULATIVE_COMPARISON_CSV, index=False)
    prompts_jsonl = write_jsonl_prompts(all_results, output_dir)

    # Plots use reverse-scored values for all comparison/difference visualisations.
    plots: Dict[str, Path] = {}
    plots["Ranking by alignment score"] = plot_ranking_bar(ranking_df, plots_dir)
    plots["Ranking by facet mean alignment"] = plot_facet_mean_bar(ranking_df, plots_dir)
    plots["OCEAN absolute-difference heatmap"] = plot_ocean_absdiff_heatmap(ranking_df, results_df, plots_dir)

    for pair, pair_df in results_df.groupby("model_pair"):
        pair_slug = slugify(pair)
        plots[f"OCEAN bars — {pair}"] = plot_pair_ocean_bars(pair_df, pair_slug, plots_dir)
        plots[f"Reverse-scored score distribution — {pair}"] = plot_pair_score_distribution(pair_df, pair_slug, plots_dir)
        plots[f"Reverse-scored difference violin — {pair}"] = plot_pair_violin(pair_df, pair_slug, plots_dir)

    excel_path = write_excel_with_plots(
        results_df=results_df,
        ranking_df=ranking_df,
        cumulative_comparison_df=cumulative_comparison_df,
        output_dir=output_dir,
        plots=plots,
        logger=logger,
    )

    expected_keys = expected_result_keys(df, selection)
    completed = expected_keys.issubset({result_unique_key(r) for r in all_results})
    mark_resume_state_completed(output_dir, completed=completed)

    logger.info("=" * 86)
    logger.info("DONE")
    logger.info(f"Run completion status: {'complete' if completed else 'incomplete — restart will offer resume'}")
    logger.info(f"Excel alignment workbook: {excel_path}")
    logger.info(f"Alignment scores CSV: {alignment_csv}")
    logger.info(f"Mean deviation CSV: {mean_dev_csv}")
    logger.info(f"Model ranking CSV: {ranking_csv}")
    logger.info(f"Current combo summary CSV: {current_comparison_csv}")
    logger.info(f"Cumulative comparison CSV: {CUMULATIVE_COMPARISON_CSV}")
    logger.info(f"Prompts JSONL: {prompts_jsonl}")

    print("\nValidation finished." if completed else "\nValidation exported, but some expected cases/pairs are still incomplete.")
    print(f"Output folder : {output_dir}")
    print(f"Excel         : {excel_path}")
    print(f"Alignment CSV : {alignment_csv}")
    print(f"Mean dev CSV  : {mean_dev_csv}")
    print(f"Ranking       : {ranking_csv}")
    print(f"Prompts JSONL : {prompts_jsonl}")
    print("\nAlignment metrics use reverse-scored values for both human and LLM responses. Questionnaire split is configurable; case-matrix and domain-stats exports are removed.")


if __name__ == "__main__":
    main()
